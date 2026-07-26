from datetime import datetime, timedelta
from fastapi import APIRouter, HTTPException, Request
from app_api.date_utils import normalize_date_yyyy_mm_dd
from fastapi.responses import StreamingResponse, JSONResponse, HTMLResponse
from database import get_db, get_company_profile
from app_api.auth import get_current_user, get_warehouse_filter_clause, get_user_assigned_warehouses
from app_api.excel_utils import fmt_money, fmt_money_dec, fmt_date
from urllib.parse import quote
from pathlib import Path
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell.cell import MergedCell

# 2 Router riêng biệt để map prefix khác nhau
router = APIRouter()      # Cho /api/reports
router_dashboard = APIRouter() # Cho /api (cho dashboard)


def _auto_fit_columns(ws, max_width: int = 40):
    for column_cells in ws.columns:
        values = []
        for cell in column_cells:
            if cell.value is not None:
                values.append(str(cell.value))
        if not values:
            continue
        first = next((cell for cell in column_cells if not isinstance(cell, MergedCell)), None)
        if not first:
            continue
        column_letter = first.column_letter
        width = min(max((len(v) for v in values), default=0) + 2, max_width)
        ws.column_dimensions[column_letter].width = width


def _style_header_row(ws, row_idx: int, start_col: int, end_col: int):
    fill = PatternFill("solid", fgColor="E2E8F0")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill


def _set_report_brand(ws, title: str):
    profile = get_company_profile()
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = profile.get("company_name") or ""
    ws["B2"] = profile.get("tax_code") or ""
    ws["D2"] = profile.get("phone") or ""
    ws["A3"] = profile.get("address") or ""
    ws["D3"] = profile.get("email") or ""


ALLOWED_TABLES = {"import_orders", "export_orders", "products", "inventory", "customers", "suppliers"}

def _count_orders_in_range(c, table_name: str, start_date: str, end_date: str, wh_clause: str, wh_params: list):
    # Validate table_name to prevent SQL injection
    if table_name not in ALLOWED_TABLES:
        return 0
    query = f"""
        SELECT COUNT(*)
        FROM {table_name}
        WHERE ((date(created_at) >= ? AND date(created_at) <= ?) OR (date(order_date) >= ? AND date(order_date) <= ?)) {wh_clause}
    """
    c.execute(query, [start_date, end_date, start_date, end_date] + wh_params)
    return c.fetchone()[0] or 0


def _get_report_warehouse_filter(user, warehouse_alias: str = 'warehouse_id'):
    """Tạo điều kiện SQL để lọc theo kho.
    - Admin/manager: không lọc (xem đầy đủ)
    - Role khác: lọc theo các warehouse được gán cho user
    """
    if not user:
        return "", []

    role = (user.get("role") or "").strip().lower()
    if role in ("admin", "manager"):
        return "", []

    wh_ids = get_user_assigned_warehouses(user.get('id'))
    if not wh_ids:
        return f" AND {warehouse_alias} IN (SELECT 1 WHERE 1=0)", []

    placeholders = ','.join(['?'] * len(wh_ids))
    return f" AND {warehouse_alias} IN ({placeholders})", wh_ids


@router.get("/export")
async def export_report(report_type: str, start_date: str = None, end_date: str = None, created_by: str = None, warehouse_id: str = None, request: Request = None):
    # Normalize date inputs for DB comparisons: DB stores *_date as YYYY-MM-DD
    start_date = normalize_date_yyyy_mm_dd(start_date)
    end_date = normalize_date_yyyy_mm_dd(end_date)

    conn = get_db()
    c = conn.cursor()

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        if report_type == "inventory":
            user = get_current_user(request)
            wh_clause, wh_params = _get_report_warehouse_filter(user, 'i.warehouse_id')
        
            ws.title = "Ton kho"
            _set_report_brand(ws, "BÁO CÁO GIÁ TRỊ TỒN KHO")

            ws["A3"] = "Ngày:"
            ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            if start_date:
                ws["A4"] = "Từ Ngày"
                ws["B4"] = start_date
            if end_date:
                ws["D4"] = "Dến ngày"
                ws["E4"] = end_date

            headers = ["STT", "Mã SP", "Tên SP", "ĐVT", "Tồn kho", "Giá vốn", "Thành tiền"]
            header_row = 6
            for idx, header in enumerate(headers, 1):
                ws.cell(row=header_row, column=idx, value=header)
            _style_header_row(ws, header_row, 1, len(headers))

            c.execute(f"""
                SELECT p.code, p.name, u.name as unit_name,
                       COALESCE(i.quantity_in_stock, 0) as stock,
                       p.cost_price,
                       ROUND(COALESCE(i.quantity_in_stock, 0) * p.cost_price, 2) as total_value
                FROM products p
                LEFT JOIN units u ON p.unit_id = u.id
                LEFT JOIN inventory i ON p.id = i.product_id
                WHERE p.is_active = 1 {wh_clause}
                ORDER BY total_value DESC
            """, wh_params)
            items = [dict(row) for row in c.fetchall()]
            row = header_row + 1
            for idx, item in enumerate(items, 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=item["code"])
                ws.cell(row=row, column=3, value=item["name"])
                ws.cell(row=row, column=4, value=item["unit_name"] or "")
                ws.cell(row=row, column=5, value=float(item["stock"] or 0))
                fmt_money_dec(ws, row, 6, item["cost_price"])
                fmt_money_dec(ws, row, 7, item["total_value"])
                row += 1

            ws.cell(row=row, column=6, value="Tổng giá trị").font = Font(bold=True)
            fmt_money_dec(ws, row, 7, sum(float(x["total_value"] or 0) for x in items))
            ws.cell(row=row, column=7).font = Font(bold=True)
            _auto_fit_columns(ws)
            filename = "bao_cao_ton_kho.xlsx"

        elif report_type == "sales":
            user = get_current_user(request)
            wh_clause, wh_params = _get_report_warehouse_filter(user, 'e.warehouse_id')
        
            ws.title = "Doanh thu"
            _set_report_brand(ws, "BÁO CÁO DOANH THU")

            ws["A3"] = "Từ Ngày"
            ws["B3"] = start_date or ""
            ws["D3"] = "Đến ngày"
            ws["E3"] = end_date or ""

            created_by_filter = ""
            created_by_params = []
            if created_by:
                created_by_filter = " AND e.created_by = ?"
                created_by_params.append(created_by)

            c.execute(f"""
                SELECT
                    e.code,
                    c.name as customer,
                    e.final_amount,
                    e.order_date,
                    COALESCE(u.full_name, u.username) as employee_name
                FROM export_orders e
                LEFT JOIN customers c ON e.customer_id = c.id
                LEFT JOIN users u ON e.created_by = u.id
                WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed'{wh_clause}{created_by_filter}
                ORDER BY e.order_date ASC
            """, (start_date, end_date, *wh_params, *created_by_params))
            orders = [dict(row) for row in c.fetchall()]

            headers = ["Mã đơn", "Nhân viên", "Khách hàng", "Thời gian", "Tổng tiền"]
            header_row = 6
            for idx, header in enumerate(headers, 1):
                ws.cell(row=header_row, column=idx, value=header)
            _style_header_row(ws, header_row, 1, len(headers))

            row = header_row + 1
            for order in orders:
                ws.cell(row=row, column=1, value=order["code"])
                ws.cell(row=row, column=2, value=order["employee_name"] or "")
                ws.cell(row=row, column=3, value=order["customer"] or "")
                ws.cell(row=row, column=4, value=fmt_date(order["order_date"]))
                fmt_money(ws, row, 5, order["final_amount"])
                row += 1

            row += 2
            ws.cell(row=row, column=1, value="Top san pham ban chay").font = Font(bold=True)
            row += 1

            c.execute(f"""
                SELECT p.name, SUM(ei.quantity_shipped) as qty, SUM(ei.total_price) as revenue
                FROM export_order_items ei
                JOIN products p ON ei.product_id = p.id
                JOIN export_orders e ON ei.order_id = e.id
                LEFT JOIN users u ON e.created_by = u.id
                WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed'{wh_clause} {created_by_filter}
                GROUP BY p.id
                ORDER BY qty DESC
                LIMIT 10
            """, [start_date, end_date, *wh_params, *created_by_params])
            top_products = [dict(row) for row in c.fetchall()]

            top_headers = ["Sản phẩm", "Số lượng", "Doanh thu"]
            for idx, header in enumerate(top_headers, 1):
                ws.cell(row=row, column=idx, value=header)
            _style_header_row(ws, row, 1, len(top_headers))

            row += 1
            for product in top_products:
                ws.cell(row=row, column=1, value=product["name"])
                ws.cell(row=row, column=2, value=float(product["qty"] or 0))
                fmt_money(ws, row, 3, product["revenue"])
                row += 1

            ws.cell(row=row + 1, column=2, value="Tổng doanh thu").font = Font(bold=True)
            fmt_money(ws, row + 1, 3, sum(float(o["final_amount"] or 0) for o in orders))
            ws.cell(row=row + 1, column=3).font = Font(bold=True)
            _auto_fit_columns(ws)
            filename = "bao_cao_doanh_thu.xlsx"

        elif report_type == "import":
            user = get_current_user(request)
            wh_clause, wh_params = _get_report_warehouse_filter(user, 'i.warehouse_id')
        
            ws.title = "Nhap hang"
            _set_report_brand(ws, "BÁO CÁO NHẬP HÀNG")

            ws["A3"] = "Từ Ngày"
            ws["B3"] = start_date or ""
            ws["D3"] = "Điến ngày"
            ws["E3"] = end_date or ""

            c.execute(f"""
                SELECT i.code, s.name as supplier, i.order_date, i.final_amount,
                       COALESCE(i.paid_amount, 0) as paid_amount,
                       COALESCE(i.final_amount, 0) - COALESCE(i.paid_amount, 0) as debt,
                       i.status,
                       COALESCE((SELECT SUM(COALESCE(NULLIF(ii.quantity_received, 0), ii.quantity_ordered, 0))
                                 FROM import_order_items ii WHERE ii.order_id = i.id), 0) as item_count
                FROM import_orders i
                LEFT JOIN suppliers s ON i.supplier_id = s.id
                WHERE i.order_date >= ? AND i.order_date <= ? 
                      AND i.status = 'completed'
                      {wh_clause}
                ORDER BY i.order_date DESC
            """, (start_date, end_date, *wh_params))
            orders = [dict(row) for row in c.fetchall()]

            headers = ["Mã phiếu", "Nhà cung cấp", "Ngày nhập", "Tổng SL", "Tổng tiền", "Dã thanh toán", "Còn nợ", "Trạng thái"]
            header_row = 6
            for idx, header in enumerate(headers, 1):
                ws.cell(row=header_row, column=idx, value=header)
            _style_header_row(ws, header_row, 1, len(headers))

            row = header_row + 1
            total_amount = 0
            total_paid = 0
            total_debt = 0
            for order in orders:
                total_amount += float(order["final_amount"] or 0)
                total_paid += float(order["paid_amount"] or 0)
                total_debt += float(order["debt"] or 0)
                ws.cell(row=row, column=1, value=order["code"])
                ws.cell(row=row, column=2, value=order["supplier"] or "")
                ws.cell(row=row, column=3, value=fmt_date(order["order_date"]))
                ws.cell(row=row, column=4, value=float(order["item_count"] or 0))
                fmt_money(ws, row, 5, order["final_amount"])
                fmt_money(ws, row, 6, order["paid_amount"])
                fmt_money(ws, row, 7, order["debt"])
                ws.cell(row=row, column=8, value=order["status"] or "")
                row += 1

            ws.cell(row=row, column=3, value="Tong cong:").font = Font(bold=True)
            fmt_money(ws, row, 5, total_amount)
            ws.cell(row=row, column=5).font = Font(bold=True)
            fmt_money(ws, row, 6, total_paid)
            ws.cell(row=row, column=6).font = Font(bold=True)
            fmt_money(ws, row, 7, total_debt)
            ws.cell(row=row, column=7).font = Font(bold=True)
            _auto_fit_columns(ws)
            filename = "bao_cao_nhap_hang.xlsx"

        elif report_type == "export_orders":
            user = get_current_user(request)
            wh_clause, wh_params = _get_report_warehouse_filter(user, 'e.warehouse_id')
        
            ws.title = "Xuat hang"
            _set_report_brand(ws, "BAO CAO DON XUAT HANG")

            ws["A3"] = "Tu ngay"
            ws["B3"] = start_date or ""
            ws["D3"] = "Den ngay"
            ws["E3"] = end_date or ""

            c.execute(f"""
                SELECT e.code, c.name as customer, e.order_date, e.final_amount,
                       COALESCE(e.discount_amount, 0) as discount,
                       COALESCE(e.paid_amount, 0) as paid_amount,
                       COALESCE(e.final_amount, 0) - COALESCE(e.paid_amount, 0) as debt,
                       e.status
                FROM export_orders e
                LEFT JOIN customers c ON e.customer_id = c.id
                WHERE e.order_date >= ? AND e.order_date <= ? {wh_clause}
                ORDER BY e.order_date DESC
            """, (start_date, end_date, *wh_params))
            orders = [dict(row) for row in c.fetchall()]

            headers = ["Ma don", "Khach hang", "Ngay ban", "Tong tien", "Giam gia", "Da thanh toan", "Con no", "TrangThai"]
            header_row = 6
            for idx, header in enumerate(headers, 1):
                ws.cell(row=header_row, column=idx, value=header)
            _style_header_row(ws, header_row, 1, len(headers))

            row = header_row + 1
            total_amount = 0
            total_discount = 0
            total_paid = 0
            total_debt = 0
            for order in orders:
                total_amount += float(order["final_amount"] or 0)
                total_discount += float(order["discount"] or 0)
                total_paid += float(order["paid_amount"] or 0)
                total_debt += float(order["debt"] or 0)
                ws.cell(row=row, column=1, value=order["code"])
                ws.cell(row=row, column=2, value=order["customer"] or "")
                ws.cell(row=row, column=3, value=fmt_date(order["order_date"]))
                fmt_money(ws, row, 4, order["final_amount"])

                fmt_money(ws, row, 5, order["discount"])
                fmt_money(ws, row, 6, order["paid_amount"])
                fmt_money(ws, row, 7, order["debt"])
                ws.cell(row=row, column=8, value=order["status"] or "")
                row += 1

            ws.cell(row=row, column=3, value="Tong cong:").font = Font(bold=True)
            fmt_money(ws, row, 4, total_amount)
            ws.cell(row=row, column=4).font = Font(bold=True)
            fmt_money(ws, row, 5, total_discount)
            ws.cell(row=row, column=5).font = Font(bold=True)
            fmt_money(ws, row, 6, total_paid)
            ws.cell(row=row, column=6).font = Font(bold=True)
            fmt_money(ws, row, 7, total_debt)
            ws.cell(row=row, column=7).font = Font(bold=True)
            _auto_fit_columns(ws)
            filename = "bao_cao_xuat_hang.xlsx"

        elif report_type == "profit":
            user = get_current_user(request)
            wh_clause, wh_params = _get_report_warehouse_filter(user, 'warehouse_id')
        
            ws.title = "Lai lo"
            _set_report_brand(ws, "BAO CAO LAI LO")

            ws["A3"] = "Tu ngay"
            ws["B3"] = start_date or ""
            ws["D3"] = "Den ngay"
            ws["E3"] = end_date or ""

            c.execute(f"""
                SELECT COALESCE(SUM(final_amount), 0) as total
                FROM import_orders
                WHERE order_date >= ? AND order_date <= ? AND status='completed' {wh_clause}
            """, (start_date, end_date, *wh_params))
            total_import = round(c.fetchone()[0] or 0, 2)

            c.execute(f"""
                SELECT COALESCE(SUM(final_amount), 0) as total
                FROM export_orders
                WHERE order_date >= ? AND order_date <= ? AND status='completed' {wh_clause}
            """, (start_date, end_date, *wh_params))
            total_export = round(c.fetchone()[0] or 0, 2)

            c.execute(f"""
                SELECT COALESCE(SUM(ei.quantity_shipped * p.cost_price), 0) as cost
                FROM export_order_items ei
                JOIN products p ON ei.product_id = p.id
                JOIN export_orders e ON ei.order_id = e.id
                WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed' {wh_clause}
            """, (start_date, end_date, *wh_params))
            cost_of_goods = round(c.fetchone()[0] or 0, 2)

            gross_profit = round(total_export - cost_of_goods, 2)
            net_profit = round(total_export - total_import, 2)

            row = 6
            ws.cell(row=row, column=1, value="Chi tieu").font = Font(bold=True)
            ws.cell(row=row, column=2, value="So tien").font = Font(bold=True)
            row += 1
            _style_header_row(ws, row, 1, 2)
            fmt_money(ws, row, 2, total_import)
            row += 1
            ws.cell(row=row, column=1, value="Tong xuat hang")
            fmt_money(ws, row, 2, total_export)
            row += 1
            ws.cell(row=row, column=1, value="Gia von hang ban")
            fmt_money(ws, row, 2, cost_of_goods)
            row += 1
            ws.cell(row=row, column=1, value="Lai gop").font = Font(bold=True)
            fmt_money(ws, row, 2, gross_profit)
            ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="Lai thuan").font = Font(bold=True)
            fmt_money(ws, row, 2, net_profit)
            ws.cell(row=row, column=2).font = Font(bold=True)

            _auto_fit_columns(ws)
            filename = "bao_cao_lai_lo.xlsx"

        else:
            raise HTTPException(status_code=400, detail="Loại báo cáo không hợp lệ")

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8\'\'{quote(filename)}"
        }
        )
    finally:
        conn.close()

@router_dashboard.get("/dashboard")
async def get_dashboard_stats(start_date: str = None, end_date: str = None, request: Request = None):
    """Dashboard stats với tùy ch"""
    # Normalize date inputs for DB comparisons
    start_date = normalize_date_yyyy_mm_dd(start_date)
    end_date = normalize_date_yyyy_mm_dd(end_date)

    user = get_current_user(request)
    wh_clause, wh_params = _get_report_warehouse_filter(user, 'warehouse_id')

    conn = get_db()
    c = conn.cursor()

    try:
        # Default: tháng hiện tại
        if not start_date:
            start_date = datetime.now().strftime("%Y-%m-01")
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")

        # Tổng sản phẩm
        c.execute("SELECT COUNT(*) FROM products WHERE is_active=1")
        total_products = c.fetchone()[0]

        # Tổng tồn kho (số lượng và giá trị)
        c.execute(
            f"SELECT COALESCE(SUM(quantity_in_stock), 0) FROM inventory WHERE 1=1 {wh_clause}",
            wh_params,
        )
        total_stock = c.fetchone()[0] or 0
        c.execute(
            f"SELECT COALESCE(SUM(quantity_in_stock * avg_cost_price), 0) FROM inventory WHERE 1=1 {wh_clause}",
            wh_params,
        )
        total_inventory_value = round(c.fetchone()[0] or 0, 2)

        # Đơn hàng nhập/xuất trong kỳ (đếm theo phiếu đã lập, ưu tiên created_at rồi fallback sang order_date)
        imports_count = _count_orders_in_range(c, "import_orders", start_date, end_date, wh_clause, wh_params)
        exports_count = _count_orders_in_range(c, "export_orders", start_date, end_date, wh_clause, wh_params)

        # Doanh thu (xuất) trong kỳ
        c.execute(
            f"""
            SELECT COALESCE(SUM(final_amount), 0) FROM export_orders
            WHERE order_date >= ? AND order_date <= ? AND status='completed' {wh_clause}
            """,
            [start_date, end_date] + wh_params,
        )
        total_revenue = round(c.fetchone()[0] or 0, 2)

        # Chi phí nhập trong kỳ
        c.execute(
            f"""
            SELECT COALESCE(SUM(final_amount), 0) FROM import_orders
            WHERE order_date >= ? AND order_date <= ? AND status='completed' {wh_clause}
            """,
            [start_date, end_date] + wh_params,
        )
        total_cost = round(c.fetchone()[0] or 0, 2)

        # Lãi lỗ = Doanh thu - Chi phí nhập
        total_profit = round(total_revenue - total_cost, 2)

        # Cảnh báo tồn thấp (đếm theo từng kho - consistent với /inventory/low-stock)
        if wh_params:
            c.execute(
                f"""
                SELECT COUNT(DISTINCT p.id || '-' || i.warehouse_id) FROM products p
                LEFT JOIN inventory i ON p.id=i.product_id
                WHERE p.is_active=1 AND p.min_stock > 0
                AND COALESCE(i.quantity_in_stock, 0) <= p.min_stock
                AND i.warehouse_id IN ({','.join(['?'] * len(wh_params))})
                """,
                wh_params,
            )
        else:
            c.execute("""
                SELECT COUNT(DISTINCT p.id || '-' || i.warehouse_id) FROM products p
                LEFT JOIN inventory i ON p.id=i.product_id
                WHERE p.is_active=1 AND p.min_stock > 0
                AND COALESCE(i.quantity_in_stock, 0) <= p.min_stock
            """)
        low_stock_count = c.fetchone()[0]

        # Biểu đồ xu hướng (7 ngày gần nhất)
        daily_chart = []
        for i in range(6, -1, -1):
            date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            c.execute(
                f"SELECT COALESCE(SUM(final_amount),0) FROM import_orders WHERE order_date LIKE ? AND status='completed' {wh_clause}",
                [f"{date}%"] + wh_params,
            )
            imp_val = round(c.fetchone()[0] or 0, 2)
            c.execute(
                f"SELECT COALESCE(SUM(final_amount),0) FROM export_orders WHERE order_date LIKE ? AND status='completed' {wh_clause}",
                [f"{date}%"] + wh_params,
            )
            exp_val = round(c.fetchone()[0] or 0, 2)
            daily_chart.append({"date": date, "import": imp_val, "export": exp_val})

        # Hoạt động gần đây
        c.execute(
            f"""
            SELECT 'import' as type, code, created_at as date, final_amount
            FROM import_orders WHERE 1=1 {wh_clause}
            ORDER BY created_at DESC LIMIT 5
            """,
            wh_params,
        )
        imports = [dict(r) for r in c.fetchall()]
        c.execute(
            f"""
            SELECT 'export' as type, code, created_at as date, final_amount
            FROM export_orders WHERE 1=1 {wh_clause}
            ORDER BY created_at DESC LIMIT 5
            """,
            wh_params,
        )
        exports = [dict(r) for r in c.fetchall()]
        recent = sorted(imports + exports, key=lambda x: x["date"], reverse=True)[:5]

        # Thông báo mới nhất (5 dòng)
        notifications = []
        
        # Lấy phiếu nhập mới nhất với tên nhân viên
        c.execute(
            f"""
            SELECT 
                'import' as type,
                io.code,
                io.created_at as date,
                io.final_amount,
                COALESCE(u.full_name, u.username) as employee_name
            FROM import_orders io
            LEFT JOIN users u ON io.created_by = u.id
            WHERE 1=1 {wh_clause.replace('warehouse_id', 'io.warehouse_id')}
            ORDER BY io.created_at DESC
            LIMIT 5
            """,
            wh_params,
        )
        import_notifications = [dict(r) for r in c.fetchall()]
        
        # Lấy phiếu xuất mới nhất với tên nhân viên
        c.execute(
            f"""
            SELECT 
                'export' as type,
                eo.code,
                eo.created_at as date,
                eo.final_amount,
                COALESCE(u.full_name, u.username) as employee_name
            FROM export_orders eo
            LEFT JOIN users u ON eo.created_by = u.id
            WHERE 1=1 {wh_clause.replace('warehouse_id', 'eo.warehouse_id')}
            ORDER BY eo.created_at DESC
            LIMIT 5
            """,
            wh_params,
        )
        export_notifications = [dict(r) for r in c.fetchall()]
        
        # Gộp và sắp xếp theo thời gian mới nhất
        all_notifications = import_notifications + export_notifications
        all_notifications.sort(key=lambda x: x["date"], reverse=True)
        notifications = all_notifications[:5]

        # Công nợ - Tổng phải trả NCC và phải thu KH
        debt_status = {
            "payable_to_suppliers": 0,  # Phải trả NCC
            "receivable_from_customers": 0  # Phải thu KH
        }
        
        try:
            # Tổng phải trả NCC (từ import_orders chưa thanh toán hết)
            c.execute(f"""
                SELECT COALESCE(SUM(final_amount - paid_amount), 0) as total
                FROM import_orders
                WHERE status = 'completed' {wh_clause.replace('warehouse_id', 'warehouse_id')}
            """, wh_params)
            debt_status["payable_to_suppliers"] = round(c.fetchone()[0] or 0, 2)
            
            # Tổng phải thu KH (từ export_orders chưa thanh toán hết)
            c.execute(f"""
                SELECT COALESCE(SUM(final_amount - paid_amount), 0) as total
                FROM export_orders
                WHERE status IN ('completed', 'shipped') {wh_clause.replace('warehouse_id', 'warehouse_id')}
            """, wh_params)
            debt_status["receivable_from_customers"] = round(c.fetchone()[0] or 0, 2)
        except Exception as e:
            # trả rỗng
            pass

        return {
            "total_products": total_products,
            "total_stock": total_stock,
            "total_inventory_value": round(total_inventory_value or 0, 2),
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "total_profit": total_profit,
            "monthly_imports": imports_count,
            "monthly_exports": exports_count,
            "low_stock_count": low_stock_count,
            "daily_chart": daily_chart,
            "recent_orders": recent,
            "notifications": notifications,
            "debt_status": debt_status,
        }

    except Exception as e:
        import traceback
        err = {
            "detail": "Dashboard query failed",
            "error_type": type(e).__name__,
            "message": str(e),
            "wh_clause": wh_clause,
            "wh_params": wh_params,
            "start_date": start_date,
            "end_date": end_date,
            "trace": traceback.format_exc(),
        }
        
        return JSONResponse(status_code=500, content=err)

    finally:
        conn.close()


@router.get("/inventory")
async def report_inventory(start_date: str = None, end_date: str = None, warehouse_id: str = None, request: Request = None):
    """Báo cáo tồn kho chi tiết theo kỳ.

    Nếu có start_date/end_date và bảng inventory_transactions tồn tại:
    - Tồn đầu kỳ: tổng quantity_after của transaction có created_at < start_date
    - Nhập trong kỳ: tổng quantity_change theo transaction_type='import' trong [start,end]
    - Xuất trong kỳ: tổng quantity_change theo transaction_type='export' trong [start,end]
    - Tồn cuối kỳ: cộng tồn đầu + nhập - xuất (theo cùng dữ liệu transaction)

    Lưu ý: hệ thống hiện dùng inventory_transactions với cột created_at.
    """
    user = get_current_user(request)
    wh_clause, wh_params = _get_report_warehouse_filter(user, 'p.warehouse_id')
    if warehouse_id:
        wh_clause = f" AND p.warehouse_id = ?"
        wh_params = [int(warehouse_id)]

    conn = get_db()
    c = conn.cursor()

    # Lấy danh sách sản phẩm trong scope kho
    c.execute(
        f"""
        SELECT p.id, p.code, p.name, u.name as unit_name,
               p.cost_price, p.selling_price,
               COALESCE(p.min_stock, 0) as min_stock,
               COALESCE(p.max_stock, 0) as max_stock
        FROM products p
        LEFT JOIN units u ON p.unit_id = u.id
        WHERE p.is_active = 1 {wh_clause}
        ORDER BY p.id
        """,
        wh_params,
    )
    products = [dict(r) for r in c.fetchall()]

    # Nếu không đủ date hoặc không có transaction => fallback tồn hiện tại
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='inventory_transactions'")
    has_tx = c.fetchone() is not None
    if not (start_date and end_date and has_tx):
        # fallback tồn hiện tại
        c.execute(
            f"""
            SELECT p.code, p.name, u.name as unit_name,
                   COALESCE(i.quantity_in_stock, 0) as stock,
                   p.cost_price,
                   p.selling_price,
                   ROUND(COALESCE(i.quantity_in_stock, 0) * p.cost_price, 2) as total_value,
                   COALESCE(p.min_stock,0) as min_stock,
                   COALESCE(p.max_stock,0) as max_stock
            FROM products p
            LEFT JOIN units u ON p.unit_id = u.id
            LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = p.warehouse_id
            WHERE p.is_active = 1 {wh_clause}
            ORDER BY p.id
            """,
            wh_params,
        )
        items = [dict(row) for row in c.fetchall()]
        total_stock = sum(x.get('stock', 0) or 0 for x in items)
        total_val = round(sum(x.get('total_value', 0) or 0 for x in items), 2)
        conn.close()
        return {
            "items": items,
            "total_stock": total_stock,
            "total_value": total_val,
            "start_date": start_date,
            "end_date": end_date,
            "mode": "current_stock_fallback",
        }

    # Tính theo transaction
    scope_wh_params = wh_params

    # Tồn đầu kỳ (lấy quantity_after của transaction gần nhất trước start_date)
    c.execute(
        f"""
        SELECT x.product_id, x.quantity_after as stock_begin
        FROM (
            SELECT it.product_id, it.quantity_after,
                   ROW_NUMBER() OVER (PARTITION BY it.product_id ORDER BY it.created_at DESC, it.id DESC) as rn
            FROM inventory_transactions it
            WHERE it.created_at < ?
              AND it.transaction_type IN ('import','import_edit','export','export_edit','adjust','import_revert','export_revert')
              AND {('1=1' if not scope_wh_params else 'it.warehouse_id IN (' + ','.join(['?']*len(scope_wh_params)) + ')')}
        ) x
        WHERE x.rn = 1
        """,
        ([start_date, *scope_wh_params] if scope_wh_params else [start_date]),
    )
    stock_begin_rows = {r["product_id"]: (r["stock_begin"] or 0) for r in c.fetchall()}

    # Nhập/xuất trong kỳ: quantity_change của import là dương, export là âm => lấy ABS cho xuất
    c.execute(
        f"""
        SELECT it.product_id,
               SUM(CASE WHEN it.transaction_type IN ('import','import_edit','import_revert') THEN it.quantity_change ELSE 0 END) as qty_import,
               SUM(CASE WHEN it.transaction_type IN ('export','export_edit','export_revert') THEN -it.quantity_change ELSE 0 END) as qty_export
        FROM inventory_transactions it
        WHERE it.created_at >= ? AND it.created_at <= ?
          AND it.transaction_type IN ('import','import_edit','export','export_edit','adjust','import_revert','export_revert')
          AND {('1=1' if not scope_wh_params else 'it.warehouse_id IN (' + ','.join(['?']*len(scope_wh_params)) + ')')}
        GROUP BY it.product_id
        """,
        [start_date, end_date, *scope_wh_params] if scope_wh_params else [start_date, end_date],
    )

    in_out = {r["product_id"]: {"qty_import": (r["qty_import"] or 0), "qty_export": (r["qty_export"] or 0)} for r in c.fetchall()}

    # Tồn cuối kỳ (lấy quantity_after của transaction gần nhất <= end_date)
    c.execute(
        f"""
        SELECT x.product_id, x.quantity_after as stock_end
        FROM (
            SELECT it.product_id, it.quantity_after,
                   ROW_NUMBER() OVER (PARTITION BY it.product_id ORDER BY it.created_at DESC, it.id DESC) as rn
            FROM inventory_transactions it
            WHERE it.created_at <= ?
              AND it.transaction_type IN ('import','import_edit','export','export_edit','adjust','import_revert','export_revert')
              AND {('1=1' if not scope_wh_params else 'it.warehouse_id IN (' + ','.join(['?']*len(scope_wh_params)) + ')')}
        ) x
        WHERE x.rn = 1
        """,
        [end_date, *scope_wh_params] if scope_wh_params else [end_date],
    )

    stock_end_rows = {r["product_id"]: (r["stock_end"] or 0) for r in c.fetchall()}

    items = []
    total_begin = 0
    total_import = 0
    total_export = 0
    total_end = 0
    total_value_end = 0

    for p in products:
        pid = p["id"]
        begin_qty = float(stock_begin_rows.get(pid, 0) or 0)
        qty_import = float(in_out.get(pid, {}).get('qty_import', 0) or 0)
        qty_export = float(in_out.get(pid, {}).get('qty_export', 0) or 0)

        # ưu tiên stock_end_rows; nếu không có thì tính begin+import-export
        computed_end = begin_qty + qty_import - qty_export
        if pid in stock_end_rows:
            end_qty = float(stock_end_rows[pid])
        else:
            end_qty = float(computed_end)

        unit_cost = float(p.get('cost_price') or 0)
        total_end_value = round(end_qty * unit_cost, 2)

        note = ""
        min_s = float(p.get('min_stock') or 0)
        max_s = float(p.get('max_stock') or 0)
        if max_s > 0 and end_qty > max_s:
            note = "Tồn trên định mức"
        elif min_s > 0 and end_qty < min_s:
            note = "Tồn dưới định mức"
        elif end_qty == 0:
            note = "Tạm hết"
        elif qty_export > 0 and qty_import == 0:
            note = note or "Đang xả hàng"

        items.append({
            "code": p.get("code"),
            "name": p.get("name"),
            "unit_name": p.get("unit_name"),
            "dvt": p.get("unit_name"),
            "stock_begin": begin_qty,
            "qty_import": qty_import,
            "qty_export": qty_export,
            "stock_end": end_qty,
            "cost_price": unit_cost,
            "selling_price": float(p.get('selling_price') or 0),
            "total_value_end": total_end_value,
            "min_stock": min_s,
            "max_stock": max_s,
            "note": note,
        })

        total_begin += begin_qty
        total_import += qty_import
        total_export += qty_export
        total_end += end_qty
        total_value_end += total_end_value

    conn.close()
    return {
        "items": items,
        "summary": {
            "total_stock_begin": total_begin,
            "total_qty_import": total_import,
            "total_qty_export": total_export,
            "total_stock_end": total_end,
            "total_value_end": round(total_value_end, 2),
        },
        "start_date": start_date,
        "end_date": end_date,
        "mode": "inventory_transactions_based",
    }


@router.get("/sales")
async def report_sales(start_date: str = None, end_date: str = None, warehouse_id: str = None, created_by: str = None, request: Request = None):
    """Báo cáo doanh thu & Top bán chạy (có thể lọc theo người lập phiếu xuất)"""
    user = get_current_user(request)
    wh_clause, wh_params = _get_report_warehouse_filter(user, 'e.warehouse_id')
    
    conn = get_db()
    c = conn.cursor()

    # Default date range
    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-01")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    # Build filters
    created_by_filter = ""
    created_by_params = []
    if created_by:
        created_by_filter = " AND e.created_by = ?"
        created_by_params.append(created_by)

    # Doanh thu
    c.execute(f"""
        SELECT
            e.code,
            c.name as customer,
            e.final_amount,
            e.order_date,
            (SELECT SUM(ei.quantity_shipped) FROM export_order_items ei WHERE ei.order_id = e.id) as item_count,
            COALESCE(e.discount_amount, 0) as discount_amount,
            COALESCE(e.paid_amount, 0) as paid_amount,
            COALESCE(u.full_name, u.username) as employee_name
        FROM export_orders e
        LEFT JOIN customers c ON e.customer_id = c.id
        LEFT JOIN users u ON e.created_by = u.id
        WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed' {wh_clause}{created_by_filter}
        ORDER BY e.order_date DESC
    """, [start_date, end_date, *wh_params, *created_by_params])
    orders = [dict(row) for row in c.fetchall()]

    # Top SP (đồng bộ cùng filter doanh thu)
    top_params = [start_date, end_date] + wh_params
    if created_by:
        top_params.append(created_by)

    c.execute(f"""
        SELECT p.name, SUM(ei.quantity_shipped) as qty, SUM(ei.total_price) as revenue
        FROM export_order_items ei
        JOIN products p ON ei.product_id = p.id
        JOIN export_orders e ON ei.order_id = e.id
        LEFT JOIN users u ON e.created_by = u.id
        WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed' {wh_clause} {created_by_filter}
        GROUP BY p.id ORDER BY qty DESC LIMIT 10
    """, top_params)
    top_products = [dict(row) for row in c.fetchall()]

    conn.close()
    return {
        "orders": orders,
        "top_products": top_products,
        "total_revenue": round(sum(float(o["final_amount"] or 0) for o in orders), 2)
    }


@router.get("/import")
async def report_imports(start_date: str = None, end_date: str = None, warehouse_id: str = None, request: Request = None):
    """Báo cáo đơn nhập hàng

    UI cột "SL SP" yêu cầu: tổng SL THỰC NHẬN = SUM(quantity_received) theo tất cả dòng.
    Chỉ lấy phiếu đã hoàn thành để tránh lệch số lượng.
    """
    user = get_current_user(request)
    wh_clause, wh_params = _get_report_warehouse_filter(user, 'i.warehouse_id')
    if warehouse_id:
        wh_clause = f" AND i.warehouse_id = ?"
        wh_params = [int(warehouse_id)]

    conn = get_db()
    c = conn.cursor()

    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-01")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    # Build item_count subquery with warehouse filter
    wh_filter_sql = ""
    if wh_clause:
        wh_filter_sql = " AND ii.warehouse_id IN (" + ",".join(['?'] * len(wh_params)) + ")"
    
    item_count_subquery = (
        "COALESCE((SELECT SUM(COALESCE(NULLIF(ii.quantity_received, 0), ii.quantity_ordered, 0))"
        " FROM import_order_items ii WHERE ii.order_id = i.id"
        + wh_filter_sql + "), 0) as item_count"
    )
    
    # Build parameter list - only add wh_params twice if warehouse filter exists
    if wh_clause:
        query_params = [start_date, end_date] + wh_params + wh_params
    else:
        query_params = [start_date, end_date]
    
    c.execute(f"""
        SELECT i.code, s.name as supplier, i.order_date, i.final_amount,
               COALESCE(i.paid_amount, 0) as paid_amount,
               COALESCE(i.final_amount, 0) - COALESCE(i.paid_amount, 0) as debt_amount,
               i.status,
               {item_count_subquery}
        FROM import_orders i
        LEFT JOIN suppliers s ON i.supplier_id = s.id
        WHERE i.order_date >= ? AND i.order_date <= ?
              AND i.status = 'completed'
              {wh_clause}
        ORDER BY i.order_date DESC
    """, query_params)
    orders = [dict(row) for row in c.fetchall()]

    conn.close()
    return {
        "orders": orders,
        "total_amount": round(sum(float(o["final_amount"] or 0) for o in orders), 2),
        "total_paid": round(sum(float(o["paid_amount"] or 0) for o in orders), 2),
        "total_debt": round(sum(float(o["debt_amount"] or 0) for o in orders), 2)
    }


@router.get("/export_orders")
async def report_exports(start_date: str = None, end_date: str = None, warehouse_id: str = None, request: Request = None):
    """Báo cáo đơn xuất hàng"""
    # Normalize date inputs for DB comparisons
    start_date = normalize_date_yyyy_mm_dd(start_date)
    end_date = normalize_date_yyyy_mm_dd(end_date)

    user = get_current_user(request)
    wh_clause, wh_params = _get_report_warehouse_filter(user, 'e.warehouse_id')
    
    conn = get_db()
    c = conn.cursor()

    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-01")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    # Build item_count subquery with warehouse filter
    wh_filter_sql = ""
    if wh_clause:
        wh_filter_sql = " AND ei.warehouse_id IN (" + ",".join(['?'] * len(wh_params)) + ")"
    
    item_count_subquery = (
        "(SELECT SUM(ei.quantity_shipped) FROM export_order_items ei"
        " WHERE ei.order_id = e.id" + wh_filter_sql + ") as item_count"
    )
    
    # Build parameter list - only add wh_params twice if warehouse filter exists
    if wh_clause:
        query_params = [start_date, end_date] + wh_params + wh_params
    else:
        query_params = [start_date, end_date]
    
    c.execute(f"""
        SELECT e.code, c.name as customer, e.order_date, e.final_amount,
               COALESCE(e.discount_amount, 0) as discount_amount,
               COALESCE(e.paid_amount, 0) as paid_amount,
               COALESCE(e.final_amount, 0) - COALESCE(e.paid_amount, 0) as debt_amount,
               {item_count_subquery},
               e.status
        FROM export_orders e
        LEFT JOIN customers c ON e.customer_id = c.id
        WHERE e.order_date >= ? AND e.order_date <= ? {wh_clause}
        ORDER BY e.order_date DESC
    """, query_params)
    orders = [dict(row) for row in c.fetchall()]

    conn.close()
    return {
        "orders": orders,
        "total_amount": round(sum(float(o["final_amount"] or 0) for o in orders), 2),
        "total_discount": round(sum(float(o["discount_amount"] or 0) for o in orders), 2),
        "total_paid": round(sum(float(o["paid_amount"] or 0) for o in orders), 2),
        "total_debt": round(sum(float(o["debt_amount"] or 0) for o in orders), 2),
        "__debug": {
            "role": (user or {}).get("role"),
            "warehouse_filter_clause": wh_clause,
            "warehouse_filter_params": wh_params,
            "start_date": start_date,
            "end_date": end_date,
            "orders_count": len(orders),
        },
    }


@router.get("/profit")
async def report_profit(start_date: str = None, end_date: str = None, warehouse_id: str = None, request: Request = None):
    """Báo cáo lãi lỗ"""
    user = get_current_user(request)
    wh_clause, wh_params = _get_report_warehouse_filter(user, 'warehouse_id')
    if warehouse_id:
        wh_clause = f" AND warehouse_id = ?"
        wh_params = [int(warehouse_id)]
    
    conn = get_db()
    c = conn.cursor()

    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-01")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    # Tổng nhập
    c.execute(f"""
        SELECT COALESCE(SUM(final_amount), 0) as total
        FROM import_orders
        WHERE order_date >= ? AND order_date <= ? AND status='completed' {wh_clause}
    """, [start_date, end_date] + wh_params)
    total_import = round(c.fetchone()[0] or 0, 2)

    # Tổng xuất
    c.execute(f"""
        SELECT COALESCE(SUM(final_amount), 0) as total
        FROM export_orders
        WHERE order_date >= ? AND order_date <= ? AND status='completed' {wh_clause}
    """, [start_date, end_date] + wh_params)
    total_export = round(c.fetchone()[0] or 0, 2)

    # Chi phí nhập (giá vốn hàng bán)
    c.execute(f"""
        SELECT COALESCE(SUM(ei.quantity_shipped * p.cost_price), 0) as cost
        FROM export_order_items ei
        JOIN products p ON ei.product_id = p.id
        JOIN export_orders e ON ei.order_id = e.id
        WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed' {wh_clause}
    """, [start_date, end_date] + wh_params)
    cost_of_goods = round(c.fetchone()[0] or 0, 2)

    gross_profit = round(total_export - cost_of_goods, 2)
    net_profit = round(total_export - total_import, 2)

    conn.close()
    return {
        "start_date": start_date,
        "end_date": end_date,
        "total_import": total_import,
        "total_export": total_export,
        "cost_of_goods": cost_of_goods,
        "gross_profit": gross_profit,
        "net_profit": net_profit
    }


# ── Fragment routes (HTML partials cho reports page) ──

def _read_fragment(name: str):
    """Đọc nội dung file fragment template từ thư mục templates/reports/"""
    path = Path(__file__).parent.parent / "templates" / "reports" / f"_{name}.html"
    return path.read_text(encoding="utf-8")


@router.get("/fragment/inventory")
async def fragment_inventory(request: Request):
    get_current_user(request)
    return HTMLResponse(content=_read_fragment("inventory"))

@router.get("/fragment/sales")
async def fragment_sales(request: Request):
    get_current_user(request)
    return HTMLResponse(content=_read_fragment("sales"))

@router.get("/fragment/import")
async def fragment_import(request: Request):
    get_current_user(request)
    return HTMLResponse(content=_read_fragment("import"))

@router.get("/fragment/export")
async def fragment_export(request: Request):
    get_current_user(request)
    return HTMLResponse(content=_read_fragment("export"))

@router.get("/fragment/top")
async def fragment_top(request: Request):
    get_current_user(request)
    return HTMLResponse(content=_read_fragment("top"))


# ── Top stats API ──

@router.get("/top")
async def report_top(start_date: str = None, end_date: str = None, warehouse_id: str = None, request: Request = None):
    """Thống kê top: sản phẩm bán chạy, tồn kho cao, nhân viên bán chạy"""
    user = get_current_user(request)
    wh_clause, wh_params = _get_report_warehouse_filter(user, 'e.warehouse_id')

    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-01")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_db()
    c = conn.cursor()

    # 1. Top sản phẩm bán chạy
    c.execute(f"""
        SELECT p.name, SUM(ei.quantity_shipped) as qty, SUM(ei.total_price) as revenue
        FROM export_order_items ei
        JOIN products p ON ei.product_id = p.id
        JOIN export_orders e ON ei.order_id = e.id
        WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed' {wh_clause}
        GROUP BY p.id ORDER BY qty DESC LIMIT 10
    """, (start_date, end_date, *wh_params))
    top_products = [dict(r) for r in c.fetchall()]

    # 2. Top tồn kho cao
    wh_clause_inv, wh_params_inv = _get_report_warehouse_filter(user, 'p.warehouse_id')
    c.execute(f"""
        SELECT p.code, p.name, u.name as unit_name,
               COALESCE(i.quantity_in_stock, 0) as stock_end,
               COALESCE(i.quantity_in_stock, 0) * p.cost_price as total_value_end
        FROM products p
        LEFT JOIN units u ON p.unit_id = u.id
        LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = p.warehouse_id
        WHERE p.is_active = 1 {wh_clause_inv}
        ORDER BY stock_end DESC LIMIT 10
    """, wh_params_inv)
    top_stock = [dict(r) for r in c.fetchall()]

    # 3. Top nhân viên bán chạy
    c.execute(f"""
        SELECT COALESCE(u.full_name, u.username) as name,
               COUNT(DISTINCT e.id) as order_count,
               COALESCE(SUM(e.final_amount), 0) as revenue
        FROM export_orders e
        LEFT JOIN users u ON e.created_by = u.id
        WHERE e.order_date >= ? AND e.order_date <= ? AND e.status='completed' {wh_clause}
        GROUP BY e.created_by ORDER BY revenue DESC LIMIT 5
    """, (start_date, end_date, *wh_params))
    top_employees = [dict(r) for r in c.fetchall()]

    conn.close()
    return {
        "top_products": top_products,
        "top_stock": top_stock,
        "top_employees": top_employees,
    }
