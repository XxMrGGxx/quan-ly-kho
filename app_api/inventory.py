from fastapi import APIRouter, HTTPException, Request, Query, File, UploadFile
from pydantic import BaseModel
from datetime import datetime
from app_api.date_utils import normalize_date_yyyy_mm_dd
from database import get_db, get_company_profile
from app_api.auth import get_current_user, check_permission, get_warehouse_filter_clause, check_warehouse_access
from app_api.excel_utils import (
    build_order_detail_template,
    style_header,
    style_body,
    set_signatures,
    fmt_money,
    excel_response as _response_excel,
    sample_wb,
)
from pathlib import Path
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell

router = APIRouter()
router_import_export = APIRouter()

LOGO_PATH = Path(__file__).resolve().parents[1] / "static" / "logo.png"


def _resolve_export_item_warehouse_id(warehouse_id, fallback_warehouse_id: int | None = None) -> int:
    if warehouse_id in (None, ""):
        warehouse_id = fallback_warehouse_id
    try:
        return int(warehouse_id)
    except (TypeError, ValueError):
        return int(fallback_warehouse_id or 1)


def _add_logo_and_brand(ws, title: str, subtitle: str = ""):
    profile = get_company_profile()
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.height = 54
            img.width = 54
            ws.add_image(img, "A2")
        except Exception:
            ws["A2"] = "LOGO"
            ws["A2"].font = Font(bold=True)
    else:
        ws["A2"] = "LOGO"
        ws["A2"].font = Font(bold=True)

    ws["C2"] = profile.get("short_name") or profile.get("company_name") or "An Tin Solution - WMS"
    line3 = subtitle or f"Ngay xuat: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["C3"] = f"{line3} | {profile.get('phone', '')}".strip(" |")
    ws["C2"].font = Font(bold=True, size=12)


def _auto_fit_columns(ws, max_width: int = 38):
    for column_cells in ws.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        first = next((cell for cell in column_cells if not isinstance(cell, MergedCell)), None)
        if not first:
            continue
        ws.column_dimensions[first.column_letter].width = min(max(len(v) for v in values) + 2, max_width)


class OrderItem(BaseModel):
    product_id: int
    warehouse_id: int = 1
    quantity: float
    unit_price: float = 0
    discount_rate: float = 0
    total_price: float = 0

class OrderCreate(BaseModel):
    partner_id: int
    order_date: str = ""
    notes: str = ""
    discount_amount: float = 0
    paid_amount: float = 0
    payment_method: str = "cash"
    items: list[OrderItem]

@router_import_export.get("/import_orders")
async def list_import_orders(request: Request):
    user = get_current_user(request)
    wh_clause, wh_params = get_warehouse_filter_clause(user, 'io.warehouse_id')
    
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute(f"""
            SELECT io.*, s.name as supplier_name, u.full_name as created_by_name
            FROM import_orders io
            LEFT JOIN suppliers s ON io.supplier_id = s.id
            LEFT JOIN users u ON io.created_by = u.id
            WHERE 1=1 {wh_clause}
            ORDER BY created_at DESC
        """, wh_params)
        items = [dict(row) for row in c.fetchall()]
        return {"items": items}
    finally:
        conn.close()


# --- NHẬP KHO ---

# IMPORTANT: Static routes MUST be defined BEFORE dynamic routes (e.g. {order_id})
# to prevent FastAPI from trying to parse "excel" / "excel-template" / "from-excel" as an integer

@router_import_export.get("/import_orders/excel")
async def export_import_orders_excel():
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            SELECT io.code, s.name as supplier_name, io.order_date, io.status, io.final_amount,
                   io.notes, w.name as warehouse_name
            FROM import_orders io
            LEFT JOIN suppliers s ON io.supplier_id = s.id
            LEFT JOIN warehouses w ON io.warehouse_id = w.id
            ORDER BY io.created_at DESC
        """)
        rows = [dict(row) for row in c.fetchall()]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phieu nhap"
        _add_logo_and_brand(ws, "DANH SACH PHIEU NHAP KHO", "Bao cao danh sach phieu nhap")

        headers = ["STT", "Ma phieu", "Nha cung cap", "Kho", "Ngay lap", "Trang thai", "Tong tien", "Ghi chu"]
        header_row = 5
        for idx, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=idx, value=header)
        style_header(ws, header_row, len(headers))

        row = header_row + 1
        for idx, item in enumerate(rows, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item["code"])
            ws.cell(row=row, column=3, value=item["supplier_name"] or "")
            ws.cell(row=row, column=4, value=item["warehouse_name"] or "")
            ws.cell(row=row, column=5, value=item["order_date"])
            ws.cell(row=row, column=6, value=item["status"])
            fmt_money(ws, row, 7, item["final_amount"])
            ws.cell(row=row, column=8, value=item["notes"] or "")
            row += 1

        style_body(ws, header_row + 1, row - 1, len(headers))
        set_signatures(ws, row + 2)
        _auto_fit_columns(ws)
        return _response_excel(wb, "danh_sach_phieu_nhap.xlsx")
    finally:
        conn.close()


@router_import_export.get("/import_orders/excel-template")
async def download_import_template():
    wb = sample_wb(
        sheet_name="NhapKho",
        headers=["STT", "Ma SP", "Ten SP", "DVT", "So luong", "Don gia", "Thanh tien", "Ghi chu"],
        example_row=[1, "SP001", "Ten san pham", "Cai", 10, 50000, 500000, ""],
    )
    _auto_fit_columns(wb.active)
    return _response_excel(wb, "mau_nhap_kho.xlsx")


# Kích thước tối đa cho file upload (10MB)
MAX_UPLOAD_SIZE = 10 * 1024 * 1024


@router_import_export.post("/import_orders/from-excel")
async def create_import_from_excel(request: Request, file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user or not check_permission(user, 'import_orders', 'create'):
        raise HTTPException(status_code=403, detail="Không có quyền tạo phiếu nhập")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận file Excel (.xlsx, .xls)")
    
    # Validate kích thước file upload
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail=f"File quá lớn. Kích thước tối đa: {MAX_UPLOAD_SIZE // (1024*1024)}MB")
    if not contents:
        raise HTTPException(status_code=400, detail="File Excel trống")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không đọc được file Excel: {str(e)}")

    conn = get_db()
    c = conn.cursor()
    try:
        items = []
        warehouse_ids = set()
        product_warehouses = {}  # ✅ KIỂM TRA: Một sản phẩm chỉ được phép ở một kho duy nhất
        
        for idx, row in enumerate(rows, 1):
            if not any(row):
                continue
            product_code = str(row[1] or row[2] or '').strip()
            product_name = str(row[2] or row[1] or '').strip()
            quantity = float(row[4] or row[3] or 0)
            unit_price = float(row[5] or 0)
            
            # Lấy warehouse_id từ cột thứ 4 (index 3) nếu có, mặc định là 1
            warehouse_id = int(row[3]) if row[3] and str(row[3]).strip().isdigit() else 1

            if not product_code and not product_name:
                continue
            if quantity <= 0:
                continue
            
            # Kiểm tra quyền truy cập warehouse
            if not check_warehouse_access(user, warehouse_id):
                raise HTTPException(status_code=403, detail=f"Dòng {idx + 1}: Không có quyền truy cập kho {warehouse_id}")
            
            warehouse_ids.add(warehouse_id)

            product_id = None
            if product_code:
                c.execute("SELECT id FROM products WHERE code = ? AND is_active = 1", (product_code,))
                found = c.fetchone()
                if found:
                    product_id = found['id']
            if not product_id and product_name:
                c.execute("SELECT id FROM products WHERE name LIKE ? AND is_active = 1 LIMIT 1", (f"%{product_name}%",))
                found = c.fetchone()
                if found:
                    product_id = found['id']
            if not product_id:
                raise HTTPException(status_code=400, detail=f"Dòng {idx + 1}: Không tìm thấy sản phẩm '{product_code or product_name}'")

            # ✅ KIỂM TRA: Một sản phẩm chỉ được phép ở một kho duy nhất
            # Kiểm tra xem sản phẩm đã có inventory trong kho khác chưa
            c.execute("""
                SELECT i.warehouse_id, w.name as warehouse_name
                FROM inventory i
                JOIN warehouses w ON i.warehouse_id = w.id
                WHERE i.product_id = ? AND i.warehouse_id != ?
                LIMIT 1
            """, (product_id, warehouse_id))
            existing_in_other = c.fetchone()
            if existing_in_other:
                c.execute("SELECT p.code, p.name FROM products p WHERE p.id = ?", (product_id,))
                prod_info = c.fetchone()
                prod_info = dict(prod_info) if prod_info else {'code': product_code, 'name': product_name}
                raise HTTPException(
                    status_code=400,
                    detail=f"Dòng {idx + 1}: Sản phẩm '{prod_info['name']}' (mã: {prod_info['code']}) đã tồn tại trong kho '{existing_in_other['warehouse_name']}'. "
                           f"Một sản phẩm chỉ được phép ở một kho duy nhất."
                )
            
            # Kiểm tra trong cùng file
            if product_id in product_warehouses:
                if product_warehouses[product_id] != warehouse_id:
                    c.execute("SELECT p.code, p.name FROM products p WHERE p.id = ?", (product_id,))
                    prod_info = c.fetchone()
                    prod_info = dict(prod_info) if prod_info else {'code': product_code, 'name': product_name}
                    raise HTTPException(
                        status_code=400,
                        detail=f"Dòng {idx + 1}: Sản phẩm '{prod_info['name']}' (mã: {prod_info['code']}) đã được thêm vào kho khác trong cùng file. "
                               f"Một sản phẩm chỉ được phép nhập vào một kho duy nhất."
                    )
            else:
                product_warehouses[product_id] = warehouse_id

            total_price = round(quantity * unit_price, 2)
            items.append({
                "product_id": product_id,
                "warehouse_id": warehouse_id,
                "quantity": quantity,
                "unit_price": unit_price,
                "total_price": total_price,
            })

        if not items:
            raise HTTPException(status_code=400, detail="Không có sản phẩm hợp lệ trong file")

        year = datetime.now().strftime("%Y")
        c.execute(f"SELECT MAX(CAST(SUBSTR(code, 7) AS INTEGER)) FROM import_orders WHERE code LIKE 'NK{year}%'")
        max_num = c.fetchone()[0]
        next_num = (max_num or 0) + 1
        code = f"NK{year}{next_num:04d}"
        
        # Sử dụng warehouse_id đầu tiên làm warehouse_id chính cho đơn hàng
        first_warehouse_id = list(warehouse_ids)[0] if warehouse_ids else 1

        order_date_norm = normalize_date_yyyy_mm_dd(datetime.now())
        c.execute("""
            INSERT INTO import_orders (code, supplier_id, warehouse_id, order_date, notes, status, created_at)
            VALUES (?, ?, ?, ?, ?, 'draft', ?)
        """, (code, 0, first_warehouse_id, order_date_norm, "Nhập từ Excel", datetime.now().isoformat()))
        order_id = c.lastrowid

        total_amount = 0
        for item in items:
            total_amount += item['total_price']
            c.execute("""
                INSERT INTO import_order_items (order_id, product_id, warehouse_id, quantity_ordered, unit_price, total_price, batch_number)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (order_id, item['product_id'], item['warehouse_id'], item['quantity'], item['unit_price'], item['total_price'], ""))

        c.execute("UPDATE import_orders SET total_amount=?, final_amount=? WHERE id=?",
                  (round(total_amount, 2), round(total_amount, 2), order_id))
        conn.commit()
        return {"id": order_id, "code": code, "message": f"Đã tạo phiếu nhập {code} với {len(items)} sản phẩm", "order_id": order_id}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý: {str(e)}")
    finally:
        conn.close()





@router_import_export.get("/import_orders/{order_id}/excel")
async def export_import_order_detail_excel(order_id: int):
    """Xuất phiếu nhập ra file Excel với template 26 dòng"""
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            SELECT io.*, s.name as supplier_name, w.name as warehouse_name,
                   u.full_name as created_by_name, u.phone as created_by_phone
            FROM import_orders io
            LEFT JOIN suppliers s ON io.supplier_id = s.id
            LEFT JOIN warehouses w ON io.warehouse_id = w.id
            LEFT JOIN users u ON io.created_by = u.id
            WHERE io.id=?
        """, (order_id,))
        order = dict(c.fetchone() or {})
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu nhập")

        c.execute("""
            SELECT ioi.*, p.code, p.name as product_name, u.name as unit_name
            FROM import_order_items ioi
            JOIN products p ON ioi.product_id = p.id
            LEFT JOIN units u ON p.unit_id = u.id
            WHERE ioi.order_id=?
        """, (order_id,))
        items = [dict(row) for row in c.fetchall()]

        # Get supplier info
        supplier_info = {"name": "", "phone": "", "address": ""}
        if order["supplier_id"]:
            c.execute("SELECT name, phone, address FROM suppliers WHERE id=?", (order["supplier_id"],))
            row_s = c.fetchone()
            if row_s:
                supplier_info = dict(row_s)

        profile = get_company_profile()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chi tiet nhap"
        
        build_order_detail_template(ws, dict(order), items, profile, supplier_info, is_import=True)
        
        return _response_excel(wb, f"phieu_nhap_{order['code']}.xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router_import_export.get("/export_orders/{order_id}/excel")
async def export_export_order_detail_excel(order_id: str):
    # allow special values for path matching (e.g. excel-batch) without 422
    try:
        order_id_int = int(order_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Không tìm thấy phiếu xuất")

    # NOTE: sử dụng order_id_int thay cho order_id (string)
    """Xuất phiếu xuất ra file Excel với template 26 dòng chuẩn"""


    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            SELECT eo.*, c.name as customer_name, w.name as warehouse_name,
                   u.full_name as created_by_name, u.phone as created_by_phone
            FROM export_orders eo
            LEFT JOIN customers c ON eo.customer_id = c.id
            LEFT JOIN warehouses w ON eo.warehouse_id = w.id
            LEFT JOIN users u ON eo.created_by = u.id
            WHERE eo.id=?
        """, (order_id_int,))
        order = dict(c.fetchone() or {})
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu xuất")

        c.execute("""
            SELECT eoi.*, p.code, p.name as product_name, u.name as unit_name
            FROM export_order_items eoi
            JOIN products p ON eoi.product_id = p.id
            LEFT JOIN units u ON p.unit_id = u.id
            WHERE eoi.order_id=?
        """, (order_id,))
        items = [dict(row) for row in c.fetchall()]

        # Get customer info
        customer_info = {"name": "", "phone": "", "address": ""}
        if order["customer_id"]:
            c.execute("SELECT name, phone, address FROM customers WHERE id=?", (order["customer_id"],))
            row_c = c.fetchone()
            if row_c:
                customer_info = dict(row_c)

        profile = get_company_profile()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chi tiet xuat"
        
        build_order_detail_template(ws, dict(order), items, profile, customer_info, is_import=False)
        
        return _response_excel(wb, f"phieu_xuat_{order['code']}.xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()



@router_import_export.post("/import_orders")
async def create_import_order(data: OrderCreate, request: Request):
    user = get_current_user(request)
    if not user or not check_permission(user, 'import_orders', 'create'):
        raise HTTPException(status_code=403, detail="Không có quyền tạo phiếu nhập")
    
    first_wh_id = data.items[0].warehouse_id if data.items else 1
    if not check_warehouse_access(user, first_wh_id):
        raise HTTPException(status_code=403, detail="Không có quyền truy cập kho này")
    
    for item in data.items:
        if not check_warehouse_access(user, item.warehouse_id):
            raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {item.warehouse_id}")
    
    # ✅ KIỂM TRA: Một sản phẩm chỉ được phép ở một kho duy nhất
    conn = get_db()
    c = conn.cursor()
    product_warehouses = {}
    for item in data.items:
        c.execute("SELECT p.code, p.name, p.warehouse_id FROM products p WHERE p.id = ? AND p.is_active = 1", (item.product_id,))
        product = c.fetchone()
        if not product:
            conn.close()
            raise HTTPException(status_code=400, detail=f"Sản phẩm ID {item.product_id} không tồn tại")
        product = dict(product)
        
        # Kiểm tra xem sản phẩm đã được gán cho kho nào chưa
        if item.product_id in product_warehouses:
            if product_warehouses[item.product_id] != item.warehouse_id:
                conn.close()
                raise HTTPException(
                    status_code=400,
                    detail=f"Sản phẩm '{product['name']}' (mã: {product['code']}) không thể nhập vào nhiều kho cùng lúc. "
                           f"Vui lòng tạo phiếu nhập riêng cho mỗi kho."
                )
        else:
            product_warehouses[item.product_id] = item.warehouse_id
    conn.close()
    
    conn = get_db()
    c = conn.cursor()
    try:
        year = datetime.now().strftime("%Y")
        c.execute(f"SELECT MAX(CAST(SUBSTR(code, 7) AS INTEGER)) FROM import_orders WHERE code LIKE 'NK{year}%'")
        max_num = c.fetchone()[0]
        next_num = (max_num or 0) + 1
        code = f"NK{year}{next_num:04d}"
        
        order_date_norm = normalize_date_yyyy_mm_dd(data.order_date) or normalize_date_yyyy_mm_dd(datetime.now())
        c.execute("""
            INSERT INTO import_orders (code, supplier_id, warehouse_id, order_date, notes, discount_amount, paid_amount, payment_method, status, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, ?)
        """, (code, data.partner_id, first_wh_id, order_date_norm, data.notes, data.discount_amount, data.paid_amount, data.payment_method, datetime.now().isoformat(), user['id']))
        order_id = c.lastrowid
        
        total_amount = 0
        for item in data.items:
            c.execute("SELECT warehouse_id FROM products WHERE id=?", (item.product_id,))
            prod_row = c.fetchone()
            if not prod_row:
                conn.close()
                raise HTTPException(status_code=400, detail=f"Sản phẩm ID {item.product_id} không tồn tại")
            product_wh_id = prod_row['warehouse_id']

            line_total = item.total_price if item.total_price else (item.quantity * item.unit_price)
            line_total = round(float(line_total or 0), 2)
            total_amount += line_total
            c.execute("""
                INSERT INTO import_order_items (order_id, product_id, warehouse_id, quantity_ordered, unit_price, discount_rate, total_price)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (order_id, item.product_id, product_wh_id, item.quantity, item.unit_price, item.discount_rate, line_total))
            
        final_amount = round(max(total_amount - float(data.discount_amount or 0), 0), 2)
        c.execute("UPDATE import_orders SET total_amount=?, final_amount=? WHERE id=?", (round(total_amount, 2), final_amount, order_id))
        conn.commit()
        return {"message": "Tạo phiếu nháp thành công"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router_import_export.get("/import_orders/{order_id}")
async def get_import_order_detail(order_id: int):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            SELECT io.*, s.name as supplier_name, w.name as warehouse_name,
                   u.full_name as created_by_name
            FROM import_orders io
            LEFT JOIN suppliers s ON io.supplier_id = s.id
            LEFT JOIN warehouses w ON io.warehouse_id = w.id
            LEFT JOIN users u ON io.created_by = u.id
            WHERE io.id=?
        """, (order_id,))
        order = c.fetchone()
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu nhập")
        c.execute("""
            SELECT ioi.*, p.code, p.name as product_name, u.name as unit_name
            FROM import_order_items ioi
            JOIN products p ON ioi.product_id = p.id
            LEFT JOIN units u ON p.unit_id = u.id
            WHERE ioi.order_id=?
        """, (order_id,))
        items = [dict(row) for row in c.fetchall()]
        return {"order": dict(order), "items": items}
    finally:
        conn.close()


@router_import_export.put("/import_orders/{order_id}")
async def update_import_order(order_id: int, data: OrderCreate, request: Request):
    user = get_current_user(request)
    if not user or not check_permission(user, 'import_orders', 'edit'):
        raise HTTPException(status_code=403, detail="Không có quyền cập nhật")
    user_id = user.get('id')
    
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM import_orders WHERE id=?", (order_id,))
        order = dict(c.fetchone() or {})
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu nhập")
        
        if not check_warehouse_access(user, order['warehouse_id']):
            raise HTTPException(status_code=403, detail="Không có quyền truy cập kho này")
        
        old_warehouse_id = order['warehouse_id']
        is_completed = order['status'] == 'completed'
        old_supplier_id = order['supplier_id']
        
        old_qtys = {}
        if is_completed:
            c.execute("SELECT product_id, warehouse_id, quantity_ordered, quantity_received FROM import_order_items WHERE order_id=?", (order_id,))
            old_items = c.fetchall()
            for item in old_items:
                wh_id = item['warehouse_id'] or old_warehouse_id
                if not check_warehouse_access(user, wh_id):
                    raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {wh_id}")
                qty = item['quantity_received'] if (item['quantity_received'] and item['quantity_received'] > 0) else item['quantity_ordered']
                key = (item['product_id'], wh_id)
                old_qtys[key] = old_qtys.get(key, 0) + qty
        
        first_wh_id = data.items[0].warehouse_id if data.items else old_warehouse_id
        if not check_warehouse_access(user, first_wh_id):
            raise HTTPException(status_code=403, detail="Không có quyền truy cập kho này")
        
        for item in data.items:
            if not check_warehouse_access(user, item.warehouse_id):
                raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {item.warehouse_id}")
        
        order_date_norm = normalize_date_yyyy_mm_dd(data.order_date) or normalize_date_yyyy_mm_dd(datetime.now())
        c.execute("""
            UPDATE import_orders
            SET supplier_id=?, warehouse_id=?, order_date=?, notes=?, discount_amount=?, paid_amount=?, payment_method=?, updated_at=?
            WHERE id=?
        """, (data.partner_id, first_wh_id, order_date_norm, data.notes, data.discount_amount, data.paid_amount, data.payment_method, datetime.now().isoformat(), order_id))
        
        c.execute("DELETE FROM import_order_items WHERE order_id=?", (order_id,))
        
        total_amount = 0
        new_qtys = {}
        for item in data.items:
            line_total = round(float(item.total_price or (item.quantity * item.unit_price) or 0), 2)
            total_amount += line_total
            c.execute("SELECT warehouse_id FROM products WHERE id=?", (item.product_id,))
            prod_row = c.fetchone()
            if not prod_row:
                raise HTTPException(status_code=400, detail=f"Sản phẩm ID {item.product_id} không tồn tại")
            product_wh_id = prod_row['warehouse_id']

            c.execute("""
                INSERT INTO import_order_items (order_id, product_id, warehouse_id, quantity_ordered, unit_price, discount_rate, total_price)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (order_id, item.product_id, product_wh_id, item.quantity, item.unit_price, item.discount_rate, line_total))

            if is_completed:
                new_qtys[(item.product_id, product_wh_id)] = new_qtys.get((item.product_id, product_wh_id), 0) + item.quantity

        if is_completed:
            all_keys = set(old_qtys) | set(new_qtys)
            for product_key in all_keys:
                old_qty = old_qtys.get(product_key, 0)
                new_qty = new_qtys.get(product_key, 0)
                diff = new_qty - old_qty
                if diff == 0:
                    continue
                prod_id, wh_id = product_key
                
                # ===== Đọc quantity_before trước khi UPDATE =====
                c.execute("SELECT quantity_in_stock, quantity_reserved FROM inventory WHERE product_id=? AND warehouse_id=?", (prod_id, wh_id))
                inv_row = c.fetchone()
                qty_before = float(inv_row['quantity_in_stock']) if inv_row else 0.0
                qty_reserved = float(inv_row['quantity_reserved']) if inv_row else 0.0
                qty_after = qty_before + diff
                
                c.execute("""
                    INSERT INTO inventory (product_id, warehouse_id, quantity_in_stock, quantity_available, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(product_id, warehouse_id)
                    DO UPDATE SET quantity_in_stock = quantity_in_stock + ?,
                                  quantity_available = (quantity_in_stock + ?) - quantity_reserved,
                                  updated_at = ?
                """, (prod_id, wh_id, diff, qty_after, datetime.now().isoformat(), diff, diff, datetime.now().isoformat()))

                c.execute("""
                    INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_type, reference_id,
                        quantity_change, quantity_before, quantity_after, notes, created_at, created_by)
                    VALUES (?, ?, 'import_edit', 'import_order', ?, ?, ?, ?, ?, ?, ?)
                """, (prod_id, wh_id, order_id, diff, qty_before, qty_after, f"Sửa phiếu nhập {order_id}", datetime.now().isoformat(), user_id))

            # Cập nhật giá nhập sản phẩm theo đơn nhập đã sửa (giữ nguyên logic cũ: ghi đè cost_price)
            updated_products = set()
            for item in data.items:
                if item.product_id in updated_products:
                    continue
                updated_products.add(item.product_id)
                # Sử dụng weighted average cost khi sửa phiếu đã hoàn thành
                if is_completed:
                    c.execute("SELECT avg_cost_price FROM inventory WHERE product_id=? AND warehouse_id=?", (item.product_id, item.warehouse_id))
                    inv_info = c.fetchone()
                    if inv_info and float(inv_info['avg_cost_price'] or 0) > 0:
                        # Giữ nguyên giá cũ (không ghi đè khi sửa)
                        pass
                    else:
                        c.execute(
                            "UPDATE products SET cost_price=? WHERE id=?",
                            (round(float(item.unit_price or 0), 2), item.product_id),
                        )
                else:
                    c.execute(
                        "UPDATE products SET cost_price=? WHERE id=?",
                        (round(float(item.unit_price or 0), 2), item.product_id),
                    )
        
        final_amount = round(max(total_amount - float(data.discount_amount or 0), 0), 2)
        c.execute("UPDATE import_orders SET total_amount=?, final_amount=? WHERE id=?", (round(total_amount, 2), final_amount, order_id))
        
        # Đồng bộ công nợ nhà cung cấp nếu đơn đã nhập hoàn thành
        if is_completed:
            from app_api.debt import _sync_supplier_debt
            if old_supplier_id:
                _sync_supplier_debt(conn, old_supplier_id)
            if data.partner_id and data.partner_id != old_supplier_id:
                _sync_supplier_debt(conn, data.partner_id)
        
        conn.commit()
        return {"id": order_id, "message": "Cập nhật phiếu nhập thành công"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# Sửa lại hàm confirm_import (dòng xử lý inventory)

@router_import_export.put("/import_orders/{order_id}/confirm")
async def confirm_import(order_id: int, request: Request):
    """Xác nhận nhập kho -> Cập nhật tồn kho thực tế"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Không xác thực được người dùng")
    user_id = user.get('id')
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM import_orders WHERE id=?", (order_id,))
        order = dict(c.fetchone() or {})
        if not order or order['status'] == 'completed':
            raise HTTPException(status_code=400, detail="Phiếu không hợp lệ hoặc đã hoàn thành")

        c.execute("SELECT * FROM import_order_items WHERE order_id=?", (order_id,))
        items = [dict(row) for row in c.fetchall()]

        for item in items:
            prod_id = item['product_id']
            # Đảm bảo dùng warehouse_id chính thức của sản phẩm từ bảng products
            c.execute("SELECT warehouse_id FROM products WHERE id=?", (prod_id,))
            p_row = c.fetchone()
            wh_id = p_row['warehouse_id'] if p_row else (item['warehouse_id'] or order['warehouse_id'])
            
            if not check_warehouse_access(user, wh_id):
                raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {wh_id}")
            qty = float(item['quantity_received'] if (item['quantity_received'] and item['quantity_received'] > 0) else item['quantity_ordered'])
            
            # ===== Đọc quantity_before trước khi UPDATE =====
            c.execute("SELECT quantity_in_stock, quantity_reserved FROM inventory WHERE product_id=? AND warehouse_id=?", (prod_id, wh_id))
            inv_row = c.fetchone()
            inv_row = dict(inv_row) if inv_row else None
            qty_before = float(inv_row['quantity_in_stock']) if inv_row else 0.0
            qty_reserved = float(inv_row['quantity_reserved']) if inv_row else 0.0
            
            qty_after = qty_before + qty
            
            # ===== UPSERT inventory =====
            c.execute("""
                INSERT INTO inventory (product_id, warehouse_id, quantity_in_stock, quantity_available, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(product_id, warehouse_id) 
                DO UPDATE SET quantity_in_stock = quantity_in_stock + ?,
                              quantity_available = (quantity_in_stock + ?) - quantity_reserved,
                              updated_at = ?
            """, (prod_id, wh_id, qty, qty_after, datetime.now().isoformat(),
                  qty, qty, datetime.now().isoformat()))
            
            # ===== Ghi transaction đầy đủ =====
            c.execute("""
                INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_type, reference_id,
                    quantity_change, quantity_before, quantity_after, created_at, created_by)
                VALUES (?, ?, 'import', 'import_order', ?, ?, ?, ?, ?, ?)
            """, (prod_id, wh_id, order_id, qty, qty_before, qty_after, datetime.now().isoformat(), user_id))

            # ===== Cập nhật cost_price theo Weighted Average Cost =====
            new_unit_price = float(item.get('unit_price', 0) or 0)
            unit_price = item.get('unit_price', 0) or 0
            if unit_price is None:
                unit_price = 0
            new_unit_price = float(unit_price)
            
            if qty_before > 0 and inv_row and float(inv_row.get('avg_cost_price') or 0) > 0:
                old_avg = float(inv_row['avg_cost_price'])
                weighted_avg = round(((qty_before * old_avg) + (qty * new_unit_price)) / (qty_before + qty), 2)
            else:
                weighted_avg = round(new_unit_price, 2)
            
            c.execute("UPDATE products SET cost_price=? WHERE id=?", (weighted_avg, prod_id))
            c.execute("""
                UPDATE inventory SET avg_cost_price=?, total_value=ROUND(quantity_in_stock * ?, 2)
                WHERE product_id=? AND warehouse_id=?
            """, (weighted_avg, weighted_avg, prod_id, wh_id))

        received_date_norm = normalize_date_yyyy_mm_dd(datetime.now())
        c.execute("UPDATE import_orders SET status='completed', received_date=? WHERE id=?", 
                  (received_date_norm, order_id))
        
        # Đồng bộ công nợ nhà cung cấp
        if order['supplier_id']:
            from app_api.debt import _sync_supplier_debt
            _sync_supplier_debt(conn, order['supplier_id'])
        
        # ===================== KIỂM TRA HẠN MỨC CÔNG NỢ =====================
        debt_warnings = []
        if order['supplier_id']:
            c.execute("SELECT ncc_debt_limit, kh_debt_limit FROM settings WHERE id=1")
            srow = c.fetchone()
            ncc_limit = float(srow['ncc_debt_limit']) if srow and srow['ncc_debt_limit'] else 0
            
            if ncc_limit > 0:
                c.execute("SELECT current_debt FROM suppliers WHERE id=?", (order['supplier_id'],))
                s_debt = c.fetchone()
                current_debt = float(s_debt['current_debt']) if s_debt else 0
                
                if current_debt > ncc_limit:
                    debt_warnings.append(
                        f"Công nợ nhà cung cấp vượt hạn mức: {current_debt:,.2f} / {ncc_limit:,.2f}"
                    )
        
        conn.commit()
        
        response = {"message": "Đã xác nhận nhập kho và cập nhật tồn kho"}
        if debt_warnings:
            response["debt_warnings"] = debt_warnings
            response["message"] = "Đã xác nhận nhập kho (có cảnh báo công nợ)"
        return response
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# --- XUẤT KHO - CHO PHÉP ÂM KHO ---

@router_import_export.get("/export_orders")
async def list_export_orders(from_date: str = None, to_date: str = None, request: Request = None):
    user = get_current_user(request)
    wh_clause, wh_params = get_warehouse_filter_clause(user, 'eo.warehouse_id')
    
    conn = get_db()
    c = conn.cursor()

    where_clauses = []
    params = []
    if from_date:
        where_clauses.append("eo.order_date >= ?")
        params.append(from_date)
    if to_date:
        where_clauses.append("eo.order_date <= ?")
        params.append(to_date)
    
    base_where = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    
    c.execute(f"""
        SELECT eo.*, c.name as customer_name, u.full_name as created_by_name
        FROM export_orders eo
        LEFT JOIN customers c ON eo.customer_id = c.id
        LEFT JOIN users u ON eo.created_by = u.id
        {base_where} {wh_clause}
        ORDER BY created_at DESC
        LIMIT 50
    """, params + wh_params)
    items = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"items": items}


# IMPORTANT: Static routes MUST be defined BEFORE dynamic routes (e.g. {order_id})
# to prevent FastAPI from trying to parse "excel-batch" as an integer

@router_import_export.get("/export_orders/excel-batch")
async def export_export_orders_excel_batch(request: Request):
    """Xuất Excel gộp nhiều phiếu xuất: mỗi phiếu 1 sheet. Nhận order_ids qua query param: order_ids=1,2,3."""

    conn = get_db()
    c = conn.cursor()
    try:
        # Frontend lưu vào sessionStorage, nhưng API không đọc được sessionStorage.
        # Backend sẽ đọc order_ids từ header/cookie/token không có sẵn trong hệ thống hiện tại,
        # nên dùng query param fallback: /excel-batch?order_ids=1,2,3
        # (Tránh thay đổi cơ chế auth hiện tại quá nhiều)
        order_ids_param = request.query_params.get('order_ids')
        if not order_ids_param:
            raise HTTPException(status_code=400, detail="Thiếu danh sách order_ids")

        try:
            order_ids = [int(x.strip()) for x in order_ids_param.split(',') if x.strip()]
        except Exception:
            raise HTTPException(status_code=400, detail="order_ids không hợp lệ")

        if not order_ids:
            raise HTTPException(status_code=400, detail="Chưa chọn phiếu xuất")

        wb = openpyxl.Workbook()
        # remove mặc định sheet
        default_ws = wb.active
        wb.remove(default_ws)

        profile = get_company_profile()

        for idx, order_id in enumerate(order_ids):
            c.execute("""
                SELECT eo.*, c.name as customer_name, w.name as warehouse_name,
                       u.full_name as created_by_name, u.phone as created_by_phone
                FROM export_orders eo
                LEFT JOIN customers c ON eo.customer_id = c.id
                LEFT JOIN warehouses w ON eo.warehouse_id = w.id
                LEFT JOIN users u ON eo.created_by = u.id
                WHERE eo.id=?
            """, (order_id,))
            order = c.fetchone()
            if not order:
                raise HTTPException(status_code=404, detail=f"Không tìm thấy phiếu xuất id={order_id}")

            c.execute("""
                SELECT eoi.*, p.code, p.name as product_name, u.name as unit_name
                FROM export_order_items eoi
                JOIN products p ON eoi.product_id = p.id
                LEFT JOIN units u ON p.unit_id = u.id
                WHERE eoi.order_id=?
            """, (order_id,))
            items = [dict(row) for row in c.fetchall()]

            customer_info = {"name": "", "phone": "", "address": ""}
            if order['customer_id']:
                c.execute("SELECT name, phone, address FROM customers WHERE id=?", (order['customer_id'],))
                row_c = c.fetchone()
                if row_c:
                    customer_info = dict(row_c)

            ws = wb.create_sheet(title=f"Phieu xuat {order['code']}")
            build_order_detail_template(ws, dict(order), items, profile, customer_info, is_import=False)

        # filename
        from datetime import date
        today_str = date.today().strftime("%d%m%Y")
        filename = f"DS_XuatKho_{today_str}.xlsx"

        return _response_excel(wb, filename)

    finally:
        conn.close()


@router_import_export.get("/export_orders/excel")
async def export_export_orders_excel():

    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            SELECT eo.code, c.name as customer_name, eo.order_date, eo.status, eo.final_amount,
                   eo.notes, w.name as warehouse_name
            FROM export_orders eo
            LEFT JOIN customers c ON eo.customer_id = c.id
            LEFT JOIN warehouses w ON eo.warehouse_id = w.id
            ORDER BY eo.created_at DESC
        """)
        rows = [dict(row) for row in c.fetchall()]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phieu xuat"
        _add_logo_and_brand(ws, "DANH SACH PHIEU XUAT KHO", "Bao cao danh sach phieu xuat")

        headers = ["STT", "Ma phieu", "Khach hang", "Kho", "Ngay lap", "Trang thai", "Tong tien", "Ghi chu"]
        header_row = 5
        for idx, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=idx, value=header)
        style_header(ws, header_row, len(headers))

        row = header_row + 1
        for idx, item in enumerate(rows, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item["code"])
            ws.cell(row=row, column=3, value=item["customer_name"] or "")
            ws.cell(row=row, column=4, value=item["warehouse_name"] or "")
            ws.cell(row=row, column=5, value=item["order_date"])
            ws.cell(row=row, column=6, value=item["status"])
            fmt_money(ws, row, 7, item["final_amount"])
            ws.cell(row=row, column=8, value=item["notes"] or "")
            row += 1

        style_body(ws, header_row + 1, row - 1, len(headers))
        set_signatures(ws, row + 2)
        _auto_fit_columns(ws)
        return _response_excel(wb, "danh_sach_phieu_xuat.xlsx")
    finally:
        conn.close()


@router_import_export.get("/export_orders/{order_id}")
async def get_export_order_detail(order_id: int):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            SELECT eo.*, c.name as customer_name, w.name as warehouse_name,
                   u.full_name as created_by_name
            FROM export_orders eo
            LEFT JOIN customers c ON eo.customer_id = c.id
            LEFT JOIN warehouses w ON eo.warehouse_id = w.id
            LEFT JOIN users u ON eo.created_by = u.id
            WHERE eo.id=?
        """, (order_id,))
        order = c.fetchone()
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu xuất")
        c.execute("""
            SELECT eoi.*, p.code, p.name as product_name, u.name as unit_name
            FROM export_order_items eoi
            JOIN products p ON eoi.product_id = p.id
            LEFT JOIN units u ON p.unit_id = u.id
            WHERE eoi.order_id=?
        """, (order_id,))
        items = [dict(row) for row in c.fetchall()]
        return {"order": dict(order), "items": items}
    finally:
        conn.close()


@router_import_export.post("/export_orders")
async def create_export_order(data: OrderCreate, request: Request):
    user = get_current_user(request)
    if not user or not check_permission(user, 'export_orders', 'create'):
        raise HTTPException(status_code=403, detail="Không có quyền tạo phiếu xuất")
    
    first_wh_id = data.items[0].warehouse_id if data.items else 1
    if not check_warehouse_access(user, first_wh_id):
        raise HTTPException(status_code=403, detail="Không có quyền truy cập kho này")
    
    for item in data.items:
        if not check_warehouse_access(user, item.warehouse_id):
            raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {item.warehouse_id}")
    
    # ✅ KIỂM TRA: Một sản phẩm chỉ được phép ở một kho duy nhất
    conn = get_db()
    c = conn.cursor()
    product_warehouses = {}
    for item in data.items:
        c.execute("SELECT p.code, p.name, p.warehouse_id FROM products p WHERE p.id = ? AND p.is_active = 1", (item.product_id,))
        product = c.fetchone()
        if not product:
            conn.close()
            raise HTTPException(status_code=400, detail=f"Sản phẩm ID {item.product_id} không tồn tại")
        product = dict(product)
        
        # Kiểm tra xem sản phẩm đã được gán cho kho nào chưa
        if item.product_id in product_warehouses:
            if product_warehouses[item.product_id] != item.warehouse_id:
                conn.close()
                raise HTTPException(
                    status_code=400,
                    detail=f"Sản phẩm '{product['name']}' (mã: {product['code']}) không thể xuất từ nhiều kho cùng lúc. "
                           f"Vui lòng tạo phiếu xuất riêng cho mỗi kho."
                )
        else:
            product_warehouses[item.product_id] = item.warehouse_id
    conn.close()
    
    conn = get_db()
    c = conn.cursor()
    try:
        year = datetime.now().strftime("%Y")
        c.execute(f"SELECT MAX(CAST(SUBSTR(code, 7) AS INTEGER)) FROM export_orders WHERE code LIKE 'XK{year}%'")
        max_num = c.fetchone()[0]
        next_num = (max_num or 0) + 1
        code = f"XK{year}{next_num:04d}"
        
        order_date_norm = normalize_date_yyyy_mm_dd(data.order_date) or normalize_date_yyyy_mm_dd(datetime.now())
        c.execute("""
            INSERT INTO export_orders (code, customer_id, warehouse_id, order_date, notes, status, paid_amount, payment_status, payment_method, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, 'draft', ?, 'paid', ?, ?, ?)
        """, (code, data.partner_id, first_wh_id, order_date_norm, data.notes,
              data.paid_amount or 0, data.payment_method or 'cash', datetime.now().isoformat(), user['id']))
        order_id = c.lastrowid
        
        total_amount = 0
        for item in data.items:
            line_total = round(item.quantity * item.unit_price, 2)
            total_amount += line_total
            line_warehouse_id = _resolve_export_item_warehouse_id(item.warehouse_id, first_wh_id)
            c.execute("""
                INSERT INTO export_order_items (order_id, product_id, warehouse_id, quantity_ordered, unit_price, total_price)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (order_id, item.product_id, line_warehouse_id, item.quantity, item.unit_price, line_total))
            
        c.execute("UPDATE export_orders SET final_amount=?, paid_amount=?, payment_status=? WHERE id=?",
                  (total_amount, total_amount, 'paid', order_id))
        conn.commit()
        return {"id": order_id, "code": code, "message": "Tạo phiếu xuất thành công"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router_import_export.put("/export_orders/{order_id}")
async def update_export_order(order_id: int, data: OrderCreate, request: Request):
    user = get_current_user(request)
    if not user or not check_permission(user, 'export_orders', 'edit'):
        raise HTTPException(status_code=403, detail="Không có quyền cập nhật")
    user_id = user.get('id')
    
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM export_orders WHERE id=?", (order_id,))
        order = dict(c.fetchone() or {})
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu xuất")
        
        if not check_warehouse_access(user, order['warehouse_id']):
            raise HTTPException(status_code=403, detail="Không có quyền truy cập kho của đơn hàng này")
        
        old_warehouse_id = order['warehouse_id']
        is_shipped = order['status'] in ['completed', 'shipped']
        old_customer_id = order['customer_id']
        
        old_qtys = {}
        if is_shipped:
            c.execute(
                "SELECT eoi.product_id, eoi.warehouse_id, eoi.quantity_ordered, eoi.quantity_shipped, p.warehouse_id as product_warehouse_id "
                "FROM export_order_items eoi "
                "JOIN products p ON eoi.product_id = p.id "
                "WHERE order_id=?",
                (order_id,)
            )
            old_items = c.fetchall()
            for item in old_items:
                wh_id = item['warehouse_id'] if item['warehouse_id'] else item['product_warehouse_id']
                if not check_warehouse_access(user, wh_id):
                    raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {wh_id}")
                # BUGFIX #3: số liệu old phải lấy theo quantity_shipped thực tế.
                qty = item['quantity_shipped'] if (item['quantity_shipped'] and item['quantity_shipped'] > 0) else item['quantity_ordered']
                key = (item['product_id'], wh_id)
                old_qtys[key] = old_qtys.get(key, 0) + qty

        
        first_wh_id = data.items[0].warehouse_id if data.items else old_warehouse_id
        if not check_warehouse_access(user, first_wh_id):
            raise HTTPException(status_code=403, detail="Không có quyền truy cập kho này")
        
        for item in data.items:
            if not check_warehouse_access(user, item.warehouse_id):
                raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {item.warehouse_id}")
        
        order_date_norm = normalize_date_yyyy_mm_dd(data.order_date) or normalize_date_yyyy_mm_dd(datetime.now())
        c.execute("""
            UPDATE export_orders
            SET customer_id=?, warehouse_id=?, order_date=?, notes=?, discount_amount=?, paid_amount=?, payment_method=?, updated_at=?
            WHERE id=?
        """, (data.partner_id, first_wh_id, order_date_norm, data.notes, data.discount_amount, data.paid_amount, data.payment_method, datetime.now().isoformat(), order_id))
        
        c.execute("DELETE FROM export_order_items WHERE order_id=?", (order_id,))
        
        total_amount = 0
        new_qtys = {}
        for item in data.items:
            line_total = round(float(item.total_price or (item.quantity * item.unit_price) or 0), 2)
            total_amount += line_total
            shipped_qty = item.quantity if is_shipped else 0
            line_warehouse_id = _resolve_export_item_warehouse_id(item.warehouse_id, first_wh_id)

            c.execute("""
                INSERT INTO export_order_items (order_id, product_id, warehouse_id, quantity_ordered, quantity_shipped, unit_price, discount_rate, total_price)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (order_id, item.product_id, line_warehouse_id, item.quantity, shipped_qty, item.unit_price, item.discount_rate, line_total))

            if is_shipped:
                new_qtys[(item.product_id, line_warehouse_id)] = new_qtys.get((item.product_id, line_warehouse_id), 0) + shipped_qty

        if is_shipped:
            all_keys = set(old_qtys) | set(new_qtys)
            for product_key in all_keys:
                old_qty = old_qtys.get(product_key, 0)
                new_qty = new_qtys.get(product_key, 0)
                diff = old_qty - new_qty
                if diff == 0:
                    continue
                prod_id, wh_id = product_key
                
                # ===== Đọc quantity_before trước khi UPDATE =====
                c.execute("SELECT quantity_in_stock, quantity_reserved FROM inventory WHERE product_id=? AND warehouse_id=?", (prod_id, wh_id))
                inv_row = c.fetchone()
                qty_before = float(inv_row['quantity_in_stock']) if inv_row else 0.0
                qty_reserved = float(inv_row['quantity_reserved']) if inv_row else 0.0
                qty_after = qty_before + diff
                
                c.execute("""
                    INSERT INTO inventory (product_id, warehouse_id, quantity_in_stock, quantity_available, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(product_id, warehouse_id) 
                    DO UPDATE SET quantity_in_stock = quantity_in_stock + ?,
                                  quantity_available = (quantity_in_stock + ?) - quantity_reserved,
                                  updated_at = ?
                """, (prod_id, wh_id, diff, qty_after, datetime.now().isoformat(),
                      diff, diff, datetime.now().isoformat()))

                c.execute("""
                    INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_type, reference_id,
                        quantity_change, quantity_before, quantity_after, notes, created_at, created_by)
                    VALUES (?, ?, 'export_edit', 'export_order', ?, ?, ?, ?, ?, ?, ?)
                """, (prod_id, wh_id, order_id, diff, qty_before, qty_after, f"Sửa phiếu xuất {order_id}", datetime.now().isoformat(), user_id))

            # Cập nhật giá bán sản phẩm theo đơn xuất đã sửa
            updated_products = set()
            for item in data.items:
                if item.product_id in updated_products:
                    continue
                updated_products.add(item.product_id)
                c.execute(
                    "UPDATE products SET selling_price=? WHERE id=?",
                    (round(float(item.unit_price or 0), 2), item.product_id),
                )

        final_amount = round(max(total_amount - float(data.discount_amount or 0), 0), 2)
        c.execute("UPDATE export_orders SET total_amount=?, final_amount=? WHERE id=?", (round(total_amount, 2), final_amount, order_id))
        
        # Đồng bộ công nợ khách hàng nếu đơn đã xuất/hoàn thành
        if is_shipped:
            from app_api.debt import _sync_customer_debt
            if old_customer_id:
                _sync_customer_debt(conn, old_customer_id)
            if data.partner_id and data.partner_id != old_customer_id:
                _sync_customer_debt(conn, data.partner_id)
        
        conn.commit()
        return {"id": order_id, "message": "Cập nhật phiếu xuất thành công"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router_import_export.delete("/import_orders/{order_id}")
async def delete_import_order(order_id: int, request: Request):
    """Xóa phiếu nhập - Chỉ admin mới có quyền"""
    current_user = get_current_user(request)
    if not current_user or not check_permission(current_user, 'import_orders', 'delete'):
        raise HTTPException(status_code=403, detail="Không có quyền xóa phiếu nhập")
    user_id = current_user.get('id')
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM import_orders WHERE id=?", (order_id,))
        order = c.fetchone()
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu nhập")
        
        # Hoàn lại tồn kho nếu phiếu nhập đã hoàn thành (completed)
        if order['status'] == 'completed':
            c.execute("SELECT product_id, warehouse_id, quantity_ordered, quantity_received FROM import_order_items WHERE order_id=?", (order_id,))
            items = c.fetchall()
            for item in items:
                prod_id = item['product_id']
                wh_id = item['warehouse_id'] or order['warehouse_id']
                qty = item['quantity_received'] if (item['quantity_received'] and item['quantity_received'] > 0) else item['quantity_ordered']
                
                # Đọc quantity trước khi UPDATE
                c.execute("SELECT quantity_in_stock, quantity_reserved FROM inventory WHERE product_id=? AND warehouse_id=?", (prod_id, wh_id))
                existing = c.fetchone()
                if existing:
                    qty_before = float(existing['quantity_in_stock'])
                    qty_reserved = float(existing['quantity_reserved']) if existing['quantity_reserved'] else 0
                    qty_after = qty_before - qty
                    
                    # Row tồn tại → UPDATE trừ đi
                    c.execute("""
                        UPDATE inventory 
                        SET quantity_in_stock = quantity_in_stock - ?,
                            quantity_available = (quantity_in_stock - ?) - quantity_reserved,
                            updated_at = ?
                        WHERE product_id=? AND warehouse_id=?
                    """, (qty, qty, datetime.now().isoformat(), prod_id, wh_id))
                else:
                    # Row không tồn tại → không làm gì (không tạo row ảo)
                    qty_before = 0
                    qty_after = 0
                
                c.execute("""
                    INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_type, reference_id,
                        quantity_change, quantity_before, quantity_after, notes, created_at, created_by)
                    VALUES (?, ?, 'import_revert', 'import_order', ?, ?, ?, ?, ?, ?, ?)
                """, (prod_id, wh_id, order_id, -qty, qty_before, qty_after, f"Xóa phiếu nhập {order['code']}", datetime.now().isoformat(), user_id))
        
        # Đồng bộ công nợ nhà cung cấp nếu xóa đơn đã nhập hoàn thành
        if order['status'] == 'completed' and order['supplier_id']:
            from app_api.debt import _sync_supplier_debt
            _sync_supplier_debt(conn, order['supplier_id'])
        
        # Xóa các items trước
        c.execute("DELETE FROM import_order_items WHERE order_id=?", (order_id,))
        c.execute("DELETE FROM import_orders WHERE id=?", (order_id,))
        conn.commit()
        return {"message": "Đã xóa phiếu nhập và cập nhật lại tồn kho thành công"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router_import_export.delete("/export_orders/{order_id}")
async def delete_export_order(order_id: int, request: Request):
    """Xóa phiếu xuất - Chỉ admin mới có quyền"""
    current_user = get_current_user(request)
    if not current_user or not check_permission(current_user, 'export_orders', 'delete'):
        raise HTTPException(status_code=403, detail="Không có quyền xóa phiếu xuất")
    user_id = current_user.get('id')
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM export_orders WHERE id=?", (order_id,))
        order = c.fetchone()
        if not order:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu xuất")
        
        # Hoàn lại tồn kho nếu phiếu xuất đã hoàn thành (completed) hoặc shipped
        if order['status'] in ['completed', 'shipped']:
            c.execute("SELECT product_id, warehouse_id, quantity_ordered, quantity_shipped FROM export_order_items WHERE order_id=?", (order_id,))
            items = c.fetchall()
            for item in items:
                prod_id = item['product_id']
                wh_id = item['warehouse_id'] or order['warehouse_id']
                qty = item['quantity_shipped'] if (item['quantity_shipped'] and item['quantity_shipped'] > 0) else item['quantity_ordered']
                
                # Đọc quantity trước khi UPDATE
                c.execute("SELECT quantity_in_stock, quantity_reserved FROM inventory WHERE product_id=? AND warehouse_id=?", (prod_id, wh_id))
                existing = c.fetchone()
                if existing:
                    qty_before = float(existing['quantity_in_stock'])
                    qty_reserved = float(existing['quantity_reserved']) if existing['quantity_reserved'] else 0
                    qty_after = qty_before + qty
                    
                    # Row tồn tại → UPDATE cộng thêm (hoàn lại tồn)
                    c.execute("""
                        UPDATE inventory 
                        SET quantity_in_stock = quantity_in_stock + ?,
                            quantity_available = (quantity_in_stock + ?) - quantity_reserved,
                            updated_at = ?
                        WHERE product_id=? AND warehouse_id=?
                    """, (qty, qty, datetime.now().isoformat(), prod_id, wh_id))
                else:
                    # Row không tồn tại → không làm gì (không tạo row ảo)
                    qty_before = 0
                    qty_after = qty
                
                c.execute("""
                    INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_type, reference_id,
                        quantity_change, quantity_before, quantity_after, notes, created_at, created_by)
                    VALUES (?, ?, 'export_revert', 'export_order', ?, ?, ?, ?, ?, ?, ?)
                """, (prod_id, wh_id, order_id, qty, qty_before, qty_after, f"Xóa phiếu xuất {order['code']}", datetime.now().isoformat(), user_id))
        
        # Đồng bộ công nợ khách hàng nếu xóa đơn đã xuất/hoàn thành
        if order['status'] in ['completed', 'shipped'] and order['customer_id']:
            from app_api.debt import _sync_customer_debt
            _sync_customer_debt(conn, order['customer_id'])
        
        # Xóa các items trước
        c.execute("DELETE FROM export_order_items WHERE order_id=?", (order_id,))
        c.execute("DELETE FROM export_orders WHERE id=?", (order_id,))
        conn.commit()
        return {"message": "Đã xóa phiếu xuất và cập nhật lại tồn kho thành công"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router_import_export.put("/export_orders/{order_id}/confirm")
async def confirm_export(order_id: int, request: Request, payload: dict = None):
    """Xác nhận xuất kho theo cấu hình cho phép xuất âm kho."""

    # payload có thể chứa confirm_negative_stock khi allow âm=ON
    # (để bắt buộc xác nhận trước khi trừ tồn xuống âm)

    conn = get_db()

    c = conn.cursor()
    try:
        c.execute("SELECT * FROM export_orders WHERE id=?", (order_id,))
        order = dict(c.fetchone() or {})
        if not order or order['status'] == 'completed':
            raise HTTPException(status_code=400, detail="Phiếu không hợp lệ hoặc đã hoàn thành")

        # Get current user for warehouse access check
        user = get_current_user(request)
        if not user:
            raise HTTPException(status_code=401, detail="Không xác thực được người dùng")
        user_id = user.get('id')

        # Load settings: allow_negative_stock
        c.execute("SELECT allow_negative_stock FROM settings WHERE id=1")
        srow = c.fetchone()
        allow_negative_stock = int(srow['allow_negative_stock']) if srow else 0

        confirm_negative_stock = False
        if payload and isinstance(payload, dict):
            confirm_negative_stock = bool(payload.get('confirm_negative_stock'))


        c.execute("SELECT * FROM export_order_items WHERE order_id=?", (order_id,))
        items = [dict(row) for row in c.fetchall()]

        # First pass: check which items will go negative
        will_go_negative = []
        for item in items:
            prod_id = item['product_id']
            wh_id = _resolve_export_item_warehouse_id(item.get('warehouse_id'), order.get('warehouse_id'))
            qty = item['quantity_shipped'] if (item['quantity_shipped'] and item['quantity_shipped'] > 0) else item['quantity_ordered']

            c.execute("""
                SELECT quantity_in_stock
                FROM inventory
                WHERE product_id=? AND warehouse_id=?
            """, (prod_id, wh_id))
            row_stock = c.fetchone()
            current_stock = float(row_stock['quantity_in_stock']) if row_stock else 0.0
            if current_stock < qty:
                will_go_negative.append({
                    "product_id": prod_id,
                    "warehouse_id": wh_id,
                    "current_stock": current_stock,
                    "needed": qty,
                })

        # Validate based on settings
        if allow_negative_stock != 1:
            # Không cho phép xuất âm => chặn nếu có item thiếu tồn
            if will_go_negative:
                details = "; ".join(
                    f"sản phẩm={x['product_id']} kho={x['warehouse_id']} tồn={x['current_stock']} cần={x['needed']}"
                    for x in will_go_negative
                )
                raise HTTPException(status_code=400, detail=f"Kho không đủ để xuất: {details}")
        else:
            # Cho phép xuất âm => chỉ bắt confirm khi thực sự có item đi âm
            if will_go_negative and not confirm_negative_stock:
                raise HTTPException(
                    status_code=400,
                    detail="Phiếu xuất sẽ làm tồn kho âm. Bắt buộc xác nhận để tiếp tục."
                )

        # Second pass: apply changes
        for item in items:
            prod_id = item['product_id']

            c.execute("SELECT id FROM products WHERE id=?", (prod_id,))
            p_row = c.fetchone()
            if not p_row:
                raise HTTPException(status_code=400, detail=f"Sản phẩm ID {prod_id} không tồn tại")

            wh_id = _resolve_export_item_warehouse_id(item.get('warehouse_id'), order.get('warehouse_id'))

            # Check warehouse access
            if not check_warehouse_access(user, wh_id):
                raise HTTPException(status_code=403, detail=f"Không có quyền truy cập kho {wh_id}")

            qty = float(item['quantity_shipped'] if (item['quantity_shipped'] and item['quantity_shipped'] > 0) else item['quantity_ordered'])

            # Cập nhật quantity_shipped để báo cáo hiển thị đúng
            c.execute("UPDATE export_order_items SET quantity_shipped = ? WHERE order_id = ? AND product_id = ?",
                      (qty, order_id, prod_id))

            # ===== Đọc quantity_before trước khi UPDATE =====
            c.execute("SELECT quantity_in_stock, quantity_reserved FROM inventory WHERE product_id=? AND warehouse_id=?", (prod_id, wh_id))
            inv_row = c.fetchone()
            qty_before = float(inv_row['quantity_in_stock']) if inv_row else 0.0
            qty_reserved = float(inv_row['quantity_reserved']) if inv_row else 0.0

            qty_after = qty_before - qty

            # ===== UPSERT inventory =====
            c.execute("""
                INSERT INTO inventory (product_id, warehouse_id, quantity_in_stock, quantity_available, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(product_id, warehouse_id)
                DO UPDATE SET quantity_in_stock = quantity_in_stock - ?,
                              quantity_available = (quantity_in_stock - ?) - quantity_reserved,
                              updated_at = ?
            """, (prod_id, wh_id, -qty, qty_after, datetime.now().isoformat(),
                  qty, qty, datetime.now().isoformat()))

            # ===== Ghi transaction đầy đủ =====
            is_negative = any(x['product_id'] == prod_id and x['warehouse_id'] == wh_id for x in will_go_negative)
            notes_text = "Xuất kho âm" if (allow_negative_stock == 1 and is_negative) else "Xuất kho"
            c.execute("""
                INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_type, reference_id,
                    quantity_change, quantity_before, quantity_after, notes, created_at, created_by)
                VALUES (?, ?, 'export', 'export_order', ?, ?, ?, ?, ?, ?, ?)
            """, (prod_id, wh_id, order_id, -qty, qty_before, qty_after, notes_text, datetime.now().isoformat(), user_id))

        # Tự động cập nhật giá bán gần nhất cho sản phẩm theo từng dòng của phiếu vừa confirm
        # (an toàn/dễ kiểm soát: giá = unit_price của export_order_items ngay tại thời điểm confirm)
        for item in items:
            unit_price = item.get('unit_price', 0) or 0
            if unit_price is None:
                continue
            prod_id = item['product_id']
            c.execute(
                "UPDATE products SET selling_price=? WHERE id=?",
                (round(float(unit_price), 2), prod_id),
            )

        shipped_date_norm = normalize_date_yyyy_mm_dd(datetime.now())
        c.execute("UPDATE export_orders SET status='completed', shipped_date=? WHERE id=?", 
                  (shipped_date_norm, order_id))
        
        # Đồng bộ công nợ khách hàng
        if order['customer_id']:
            from app_api.debt import _sync_customer_debt
            _sync_customer_debt(conn, order['customer_id'])
        
        # ===================== KIỂM TRA HẠN MỨC CÔNG NỢ =====================
        debt_warnings = []
        if order['customer_id']:
            c.execute("SELECT ncc_debt_limit, kh_debt_limit FROM settings WHERE id=1")
            srow = c.fetchone()
            kh_limit = float(srow['kh_debt_limit']) if srow and srow['kh_debt_limit'] else 0
            
            if kh_limit > 0:
                c.execute("SELECT current_debt FROM customers WHERE id=?", (order['customer_id'],))
                c_debt = c.fetchone()
                current_debt = float(c_debt['current_debt']) if c_debt else 0
                
                if current_debt > kh_limit:
                    debt_warnings.append(
                        f"Công nợ khách hàng vượt hạn mức: {current_debt:,.2f} / {kh_limit:,.2f}"
                    )
        
        conn.commit()
        
        response = {}
        if will_go_negative:
            response["message"] = f"Đã xuất kho thành công ({len(will_go_negative)} sản phẩm tồn kho âm)"
        else:
            response["message"] = "Đã xuất kho thành công"
        
        if debt_warnings:
            response["debt_warnings"] = debt_warnings
            response["message"] = "Đã xác nhận xuất kho (có cảnh báo công nợ)"
        
        return response
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# --- THÊM API KIỂM TRA TỒN KHO ÂM ---

@router.get("/negative-stock")
async def get_negative_stock(request: Request, warehouse_id: int = None):
    """Lấy danh sách sản phẩm có tồn kho âm"""
    user = get_current_user(request)
    # filter theo warehouse_id trên alias i trong CTE (inventory_agg)
    wh_clause, wh_params = get_warehouse_filter_clause(user, 'i.warehouse_id')

    # Nếu có warehouse_id từ query param thì thêm vào điều kiện lọc
    if warehouse_id is not None:
        wh_clause += " AND i.warehouse_id = ?"
        wh_params.append(warehouse_id)
    
    conn = get_db()
    c = conn.cursor()
    c.execute(f"""
        SELECT p.id as product_id, p.code, p.name, i.quantity_in_stock, i.warehouse_id, w.name as warehouse_name
        FROM inventory i
        JOIN products p ON i.product_id = p.id AND p.warehouse_id = i.warehouse_id
        JOIN warehouses w ON i.warehouse_id = w.id
        WHERE i.quantity_in_stock < 0 {wh_clause}
        ORDER BY i.quantity_in_stock ASC
    """, wh_params)
    items = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"items": items, "count": len(items), "message": "Danh sách sản phẩm tồn kho âm"}

@router.post("/adjust-stock")
async def adjust_stock(data: dict, request: Request):
    """Điều chỉnh tồn kho thủ công"""
    user = get_current_user(request)
    if not user or not check_permission(user, 'inventory', 'edit'):
        raise HTTPException(status_code=403, detail="Không có quyền điều chỉnh tồn kho")
    user_id = user.get('id')
    
    product_id = data.get('product_id')
    warehouse_id = data.get('warehouse_id', 1)
    new_quantity = float(data.get('new_quantity', 0))
    reason = data.get('reason', 'Điều chỉnh thủ công')
    
    if not check_warehouse_access(user, warehouse_id):
        raise HTTPException(status_code=403, detail="Không có quyền truy cập kho này")
    
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT quantity_in_stock, quantity_reserved FROM inventory WHERE product_id=? AND warehouse_id=?", 
                  (product_id, warehouse_id))
        current = c.fetchone()
        qty_before = float(current['quantity_in_stock']) if current else 0.0
        qty_reserved = float(current['quantity_reserved']) if current and current['quantity_reserved'] else 0.0
        qty_after = new_quantity
        change = qty_after - qty_before
        
        c.execute("""
            INSERT INTO inventory (product_id, warehouse_id, quantity_in_stock, quantity_available, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(product_id, warehouse_id) 
            DO UPDATE SET quantity_in_stock = ?,
                          quantity_available = ? - quantity_reserved,
                          updated_at = ?
        """, (product_id, warehouse_id, new_quantity, qty_after - qty_reserved, datetime.now().isoformat(),
              new_quantity, qty_after - qty_reserved, datetime.now().isoformat()))
        
        # Tạo reference_id duy nhất cho giao dịch này
        c.execute("SELECT last_insert_rowid()")
        reference_id = c.fetchone()[0]
        
        c.execute("""
            INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_type, reference_id,
                quantity_change, quantity_before, quantity_after, notes, created_at, created_by)
            VALUES (?, ?, 'adjust', 'adjustment', ?, ?, ?, ?, ?, ?, ?)
        """, (product_id, warehouse_id, reference_id, change, qty_before, qty_after, reason, datetime.now().isoformat(), user_id))
        
        conn.commit()
        return {"message": f"Đã điều chỉnh tồn kho từ {qty_before} thành {new_quantity}"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/stock")
async def get_inventory_stock(request: Request, warehouse_id: int = None, search: str = Query(""), limit: int = Query(50), offset: int = Query(0)):
    """Lấy danh sách tồn kho hiện tại - mỗi sản phẩm/kho là 1 dòng riêng"""
    user = get_current_user(request)
    wh_clause, wh_params = get_warehouse_filter_clause(user, 'i.warehouse_id')
    
    # Nếu có warehouse_id từ query param thì thêm vào điều kiện lọc
    if warehouse_id is not None:
        wh_clause += " AND i.warehouse_id = ?"
        wh_params.append(warehouse_id)

    # Tìm kiếm theo tên hoặc mã sản phẩm
    search_clause = ""
    search_params = []
    if search and search.strip():
        search_clause = " AND (p.name LIKE ? OR p.code LIKE ?)"
        term = f"%{search.strip()}%"
        search_params = [term, term]
    
    conn = get_db()
    c = conn.cursor()

    # Count query
    count_params = wh_params[:] + search_params
    c.execute(f"""
        SELECT COUNT(*)
        FROM products p
        LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = p.warehouse_id
        WHERE p.is_active = 1 {wh_clause} {search_clause}
    """, count_params)
    total = c.fetchone()[0]

    # Data query with pagination
    params = wh_params[:] + search_params + [limit, offset]
    c.execute(f"""
        SELECT 
            p.id as product_id,
            p.code as product_code,
            p.name as product_name,
            COALESCE(i.quantity_in_stock, 0) as quantity_in_stock,
            p.warehouse_id as warehouse_id,
            COALESCE(w.name, 'Kho chính') as warehouse_name,
            p.min_stock
        FROM products p
        LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = p.warehouse_id
        LEFT JOIN warehouses w ON p.warehouse_id = w.id
        WHERE p.is_active = 1 {wh_clause} {search_clause}
        ORDER BY p.name, w.name
        LIMIT ? OFFSET ?
    """, params)
    items = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@router.get("/audit")
async def audit_inventory(request: Request):
    """Kiểm tra tồn kho chi tiết: xem giao dịch nào đã tạo ra số tồn hiện tại"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Không xác thực được người dùng")
    
    conn = get_db()
    c = conn.cursor()
    try:
        # Lấy tổng tồn kho
        c.execute("""
            WITH inventory_agg AS (
                SELECT product_id, warehouse_id, SUM(quantity_in_stock) as quantity_in_stock
                FROM inventory
                GROUP BY product_id, warehouse_id
            )
            SELECT p.id, p.code, p.name, 
                   COALESCE(i.quantity_in_stock, 0) as quantity_in_stock,
                   COALESCE(i.warehouse_id, 1) as warehouse_id,
                   COALESCE(w.name, 'Kho chính') as warehouse_name
            FROM products p
            LEFT JOIN inventory_agg i ON p.id = i.product_id
            LEFT JOIN warehouses w ON i.warehouse_id = w.id
            WHERE p.is_active = 1 AND COALESCE(i.quantity_in_stock, 0) != 0
            ORDER BY p.name
        """)
        inventory_items = [dict(row) for row in c.fetchall()]
        
        # Tính tổng tồn kho từ các giao dịch
        c.execute("""
            SELECT SUM(CASE 
                WHEN transaction_type IN ('import', 'import_edit') THEN quantity_change
                WHEN transaction_type IN ('export', 'export_edit') THEN quantity_change
                WHEN transaction_type = 'import_revert' THEN quantity_change
                WHEN transaction_type = 'export_revert' THEN quantity_change
                WHEN transaction_type = 'adjust' THEN 0
                ELSE 0
            END) as calculated_stock
            FROM inventory_transactions
        """)
        calc_row = c.fetchone()
        calculated_total = float(calc_row['calculated_stock']) if calc_row and calc_row['calculated_stock'] else 0
        
        # Lấy tổng từ bảng inventory
        c.execute("SELECT SUM(quantity_in_stock) as reported_total FROM inventory")
        report_row = c.fetchone()
        reported_total = float(report_row['reported_total']) if report_row and report_row['reported_total'] else 0
        
        # Số lượng import_orders đã completed
        c.execute("SELECT COUNT(*) as cnt FROM import_orders WHERE status='completed'")
        completed_imports = c.fetchone()['cnt']
        
        # Số lượng export_orders đã completed
        c.execute("SELECT COUNT(*) as cnt FROM export_orders WHERE status='completed'")
        completed_exports = c.fetchone()['cnt']
        
        return {
            "inventory_items": inventory_items,
            "inventory_count": len(inventory_items),
            "reported_total_inventory": reported_total,
            "calculated_from_transactions": calculated_total,
            "difference": reported_total - calculated_total,
            "completed_import_orders": completed_imports,
            "completed_export_orders": completed_exports,
            "note": "Nếu reported_total khác calculated_from_transactions, tồn kho đã bị lệch do lỗi cũ"
        }
    finally:
        conn.close()


@router.post("/rebuild")
async def rebuild_inventory(request: Request):
    """Xóa toàn bộ dữ liệu tồn kho và tính toán lại từ đầu dựa trên các giao dịch đã hoàn thành.
    Chỉ admin mới có quyền này."""
    user = get_current_user(request)
    if not user or user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Chỉ admin mới có quyền rebuild tồn kho")
    
    conn = get_db()
    c = conn.cursor()
    try:
        # 1. Xóa toàn bộ dữ liệu inventory cũ
        c.execute("DELETE FROM inventory")
        c.execute("DELETE FROM inventory_transactions")
        c.execute("DELETE FROM product_locations")
        
        # 2. Tính toán lại tồn kho từ các phiếu nhập đã completed
        c.execute("""
            SELECT ioi.product_id, 
                   COALESCE(ioi.warehouse_id, io.warehouse_id) as warehouse_id,
                   SUM(COALESCE(ioi.quantity_received, ioi.quantity_ordered)) as total_imported
            FROM import_order_items ioi
            JOIN import_orders io ON ioi.order_id = io.id
            WHERE io.status = 'completed'
            GROUP BY ioi.product_id, COALESCE(ioi.warehouse_id, io.warehouse_id)
        """)
        import_rows = c.fetchall()
        
        for row in import_rows:
            product_id = row['product_id']
            warehouse_id = row['warehouse_id']
            total_qty = float(row['total_imported'])
            
            c.execute("""
                INSERT INTO inventory (product_id, warehouse_id, quantity_in_stock, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(product_id, warehouse_id) 
                DO UPDATE SET quantity_in_stock = quantity_in_stock + ?,
                              updated_at = ?
            """, (product_id, warehouse_id, total_qty, datetime.now().isoformat(), 
                  total_qty, datetime.now().isoformat()))
            
            c.execute("""
                INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_id, quantity_change, notes, created_at)
                VALUES (?, ?, 'import', 0, ?, 'Rebuild từ phiếu nhập', ?)
            """, (product_id, warehouse_id, total_qty, datetime.now().isoformat()))
        
        # 3. Trừ tồn kho từ các phiếu xuất đã completed
        c.execute("""
            SELECT eoi.product_id, 
                   COALESCE(eoi.warehouse_id, eo.warehouse_id) as warehouse_id,
                   SUM(COALESCE(eoi.quantity_shipped, eoi.quantity_ordered)) as total_exported
            FROM export_order_items eoi
            JOIN export_orders eo ON eoi.order_id = eo.id
            WHERE eo.status = 'completed'
            GROUP BY eoi.product_id, COALESCE(eoi.warehouse_id, eo.warehouse_id)
        """)
        export_rows = c.fetchall()
        
        for row in export_rows:
            product_id = row['product_id']
            warehouse_id = row['warehouse_id']
            total_qty = float(row['total_exported'])
            
            # Chỉ UPDATE nếu row tồn tại
            c.execute("SELECT quantity_in_stock FROM inventory WHERE product_id=? AND warehouse_id=?", 
                      (product_id, warehouse_id))
            existing = c.fetchone()
            if existing:
                current_qty = float(existing['quantity_in_stock'])
                new_qty = max(current_qty - total_qty, 0)  # Không cho âm
                c.execute("""
                    UPDATE inventory SET quantity_in_stock=?, updated_at=?
                    WHERE product_id=? AND warehouse_id=?
                """, (new_qty, datetime.now().isoformat(), product_id, warehouse_id))
            # Nếu không có row thì không làm gì (tránh tạo tồn kho ảo)
            
            c.execute("""
                INSERT INTO inventory_transactions (product_id, warehouse_id, transaction_type, reference_id, quantity_change, notes, created_at)
                VALUES (?, ?, 'export', 0, ?, 'Rebuild từ phiếu xuất', ?)
            """, (product_id, warehouse_id, -total_qty, datetime.now().isoformat()))
        
        # 4. Xóa các row inventory có quantity = 0
        c.execute("DELETE FROM inventory WHERE quantity_in_stock = 0 OR quantity_in_stock IS NULL")
        
        conn.commit()
        
        # Đếm lại
        c.execute("SELECT COUNT(*) as cnt FROM inventory")
        cnt = c.fetchone()['cnt']
        c.execute("SELECT SUM(quantity_in_stock) as total FROM inventory")
        total = float(c.fetchone()['total'] or 0)
        
        return {
            "message": f"Đã rebuild tồn kho thành công! Còn {cnt} sản phẩm có tồn kho, tổng số lượng: {total:,.2f}",
            "inventory_items_count": cnt,
            "total_quantity": total
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/low-stock")
async def get_low_stock(request: Request, warehouse_id: int = None):
    """Lấy danh sách sản phẩm tồn kho sắp hết (quantity_in_stock <= min_stock) theo từng kho"""
    user = get_current_user(request)
    wh_clause, wh_params = get_warehouse_filter_clause(user, 'i.warehouse_id')
    
    # Nếu có warehouse_id từ query param thì thêm vào điều kiện lọc
    if warehouse_id is not None:
        wh_clause += " AND i.warehouse_id = ?"
        wh_params.append(warehouse_id)
    
    conn = get_db()
    c = conn.cursor()
    c.execute(f"""
        SELECT 
            p.id as product_id,
            p.code as product_code,
            p.name as product_name,
            COALESCE(i.quantity_in_stock, 0) as quantity_in_stock,
            p.warehouse_id as warehouse_id,
            COALESCE(w.name, 'Kho chính') as warehouse_name,
            p.min_stock
        FROM products p
        LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = p.warehouse_id
        LEFT JOIN warehouses w ON p.warehouse_id = w.id
        WHERE p.is_active = 1 
          AND p.min_stock > 0
          AND COALESCE(i.quantity_in_stock, 0) <= p.min_stock {wh_clause}
        ORDER BY p.name, w.name
    """, wh_params)
    items = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"items": items, "count": len(items)}
