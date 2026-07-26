"""
Data Management API - Backup, Restore, Import/Export, Clear Data
"""
from fastapi import APIRouter, HTTPException, Request
from pydantic import BaseModel
from datetime import datetime
from typing import Optional
from database import get_db, hash_password
from app_api.auth import get_current_user
from app_api.excel_utils import excel_response, sample_wb, load_workbook_from_base64
from path_utils import get_app_dir
import io
import os
import shutil
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse

router = APIRouter()

BACKUP_DIR = get_app_dir() / "data" / "backups"

# ========== SAFETY KEY ==========
# Khóa an toàn "Antin@2025" - bảo vệ các chức năng nhạy cảm
SAFETY_KEY_HASH = hash_password("Antin@2025")

def verify_safety_key(key: str):
    """Xác thực khóa an toàn, trả về True nếu đúng"""
    if not key:
        raise HTTPException(status_code=403, detail="Vui lòng nhập khóa an toàn")
    if hash_password(key) != SAFETY_KEY_HASH:
        raise HTTPException(status_code=403, detail="Sai khóa an toàn")
    return True

def _ensure_backup_dir():
    os.makedirs(str(BACKUP_DIR), exist_ok=True)

def _admin_only(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    if user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Chỉ admin mới có quyền")
    return user

# ========== BACKUP & RESTORE ==========

@router.post("/data/backup")
async def backup_database(request: Request):
    """Tạo file backup database"""
    _admin_only(request)
    _ensure_backup_dir()
    
    db_path = str(get_app_dir() / "data" / "warehouse.db")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Không tìm thấy database")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"warehouse_backup_{timestamp}.db"
    backup_path = str(BACKUP_DIR / backup_filename)
    
    try:
        shutil.copy2(db_path, backup_path)
        file_size = os.path.getsize(backup_path)
        return {
            "success": True,
            "message": f"Đã sao lưu database thành công",
            "filename": backup_filename,
            "file_size": file_size,
            "file_size_display": _format_size(file_size),
            "created_at": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi sao lưu: {str(e)}")


@router.get("/data/backups")
async def list_backups(request: Request):
    """Danh sách các file backup"""
    _admin_only(request)
    _ensure_backup_dir()
    
    backups = []
    try:
        files = sorted(os.listdir(str(BACKUP_DIR)), reverse=True)
        for f in files:
            if f.endswith(".db"):
                fpath = str(BACKUP_DIR / f)
                stat = os.stat(fpath)
                backups.append({
                    "filename": f,
                    "file_size": stat.st_size,
                    "file_size_display": _format_size(stat.st_size),
                    "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
                })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi đọc danh sách backup: {str(e)}")
    
    return {"items": backups, "total": len(backups)}


@router.post("/data/restore")
async def restore_database(payload: dict, request: Request):
    """Phục hồi database từ file backup"""
    _admin_only(request)
    
    filename = payload.get("filename", "")
    if not filename:
        raise HTTPException(status_code=400, detail="Thiếu tên file backup")
    
    # Security: prevent path traversal
    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Tên file không hợp lệ")
    
    backup_path = str(BACKUP_DIR / filename)
    if not os.path.exists(backup_path):
        raise HTTPException(status_code=404, detail="File backup không tồn tại")
    
    db_path = str(get_app_dir() / "data" / "warehouse.db")
    
    try:
        # Backup current database before restore (just in case)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pre_restore_backup = str(BACKUP_DIR / f"pre_restore_{timestamp}.db")
        if os.path.exists(db_path):
            shutil.copy2(db_path, pre_restore_backup)
        
        # Restore
        shutil.copy2(backup_path, db_path)
        return {
            "success": True,
            "message": "Đã phục hồi database thành công. Vui lòng đăng nhập lại."
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi phục hồi: {str(e)}")


@router.delete("/data/backups/{filename}")
async def delete_backup(filename: str, request: Request):
    """Xóa file backup"""
    _admin_only(request)
    
    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Tên file không hợp lệ")
    
    backup_path = str(BACKUP_DIR / filename)
    if not os.path.exists(backup_path):
        raise HTTPException(status_code=404, detail="File backup không tồn tại")
    
    try:
        os.remove(backup_path)
        return {"success": True, "message": "Đã xóa file backup"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xóa: {str(e)}")


@router.get("/data/backups/{filename}/download")
async def download_backup(filename: str, request: Request):
    """Tải file backup về máy"""
    _admin_only(request)
    
    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Tên file không hợp lệ")
    
    backup_path = str(BACKUP_DIR / filename)
    if not os.path.exists(backup_path):
        raise HTTPException(status_code=404, detail="File backup không tồn tại")
    
    def iterfile():
        with open(backup_path, "rb") as f:
            yield from f
    
    return StreamingResponse(
        iterfile(),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ========== EXPORT DANH MỤC EXCEL (YÊU CẦU KHÓA AN TOÀN) ==========

@router.get("/data/export/{data_type}")
async def export_data(data_type: str, request: Request, safety_key: Optional[str] = None):
    """Export danh mục ra Excel (yêu cầu khóa an toàn)
    data_type: products, customers, suppliers
    """
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    
    # Yêu cầu khóa an toàn
    verify_safety_key(safety_key)
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        if data_type == "products":
            return _export_products(c)
        elif data_type == "customers":
            return _export_customers(c)
        elif data_type == "suppliers":
            return _export_suppliers(c)
        else:
            raise HTTPException(status_code=400, detail="Loại dữ liệu không hợp lệ. Chấp nhận: products, customers, suppliers")
    finally:
        conn.close()


def _export_products(c):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hàng hóa"
    
    headers = ["Mã SP", "Barcode", "Tên hàng hóa", "Danh mục", "ĐVT", "Kho", "Giá vốn", "Giá bán", "Tồn kho", "Tồn tối thiểu", "% Chiết khấu"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    
    c.execute("""
        SELECT p.code, p.barcode, p.name, 
               COALESCE(c.name, '') as category_name,
               COALESCE(u.name, '') as unit_name,
               COALESCE(w.name, '') as warehouse_name,
               p.cost_price, p.selling_price,
               COALESCE(i.quantity_in_stock, 0) as stock,
               p.min_stock, p.discount_rate
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        LEFT JOIN units u ON p.unit_id = u.id
        LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = p.warehouse_id
        LEFT JOIN warehouses w ON p.warehouse_id = w.id
        WHERE p.is_active = 1
        ORDER BY p.code
    """)
    for row in c.fetchall():
        ws.append([row['code'], row['barcode'], row['name'], row['category_name'],
                   row['unit_name'], row['warehouse_name'],
                   row['cost_price'], row['selling_price'],
                   row['stock'], row['min_stock'], row['discount_rate']])
    
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    return excel_response(wb, "danh_sach_hang_hoa.xlsx")


def _export_customers(c):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Khách hàng"
    
    headers = ["Mã KH", "Tên khách hàng", "Người liên hệ", "SĐT", "Email", "Địa chỉ", "Khu vực", "Mã số thuế", "Công nợ"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    
    c.execute("""
        SELECT code, name, contact_person, phone, email, address, city, tax_code, current_debt
        FROM customers WHERE is_active = 1 ORDER BY code
    """)
    for row in c.fetchall():
        ws.append([row['code'], row['name'], row['contact_person'], row['phone'],
                   row['email'], row['address'], row['city'], row['tax_code'], row['current_debt']])
    
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    return excel_response(wb, "danh_sach_khach_hang.xlsx")


def _export_suppliers(c):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhà cung cấp"
    
    headers = ["Mã NCC", "Tên nhà cung cấp", "Người liên hệ", "SĐT", "Email", "Địa chỉ", "Khu vực", "Mã số thuế", "Công nợ"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    
    c.execute("""
        SELECT code, name, contact_person, phone, email, address, city, tax_code, current_debt
        FROM suppliers WHERE is_active = 1 ORDER BY code
    """)
    for row in c.fetchall():
        ws.append([row['code'], row['name'], row['contact_person'], row['phone'],
                   row['email'], row['address'], row['city'], row['tax_code'], row['current_debt']])
    
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    return excel_response(wb, "danh_sach_nha_cung_cap.xlsx")


# ========== IMPORT TEMPLATES (YÊU CẦU KHÓA AN TOÀN) ==========

@router.get("/data/import-template/{data_type}")
async def download_import_template(data_type: str, request: Request, safety_key: Optional[str] = None):
    """Download template Excel mẫu cho import (yêu cầu khóa an toàn)"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    
    # Yêu cầu khóa an toàn
    verify_safety_key(safety_key)
    
    if data_type == "products":
        wb = sample_wb("Hàng hóa", 
            ["Mã SP", "Barcode", "Tên hàng hóa", "Danh mục", "ĐVT", "Kho", "Giá vốn", "Giá bán", "Tồn kho ban đầu", "Tồn tối thiểu", "% Chiết khấu"],
            ["", "893xxxx", "Tên sản phẩm", "Điện tử", "Cái", "Kho chính", 10000, 12000, 0, 5, 0])
        return excel_response(wb, "mau_import_hang_hoa.xlsx")
    elif data_type == "customers":
        wb = sample_wb("Khách hàng",
            ["name", "contact_person", "phone", "email", "address", "city", "tax_code"],
            ["Công ty ABC", "Nguyễn Văn A", "0901234567", "abc@gmail.com", "Quận 1, TP.HCM", "HCM", "0312345678"])
        return excel_response(wb, "mau_import_khach_hang.xlsx")
    elif data_type == "suppliers":
        wb = sample_wb("Nhà cung cấp",
            ["name", "contact_person", "phone", "email", "address", "city", "tax_code"],
            ["Nhà cung cấp A", "Trần Văn B", "0909123456", "ncc@gmail.com", "Quận 3, TP.HCM", "HCM", "0319988776"])
        return excel_response(wb, "mau_import_nha_cung_cap.xlsx")
    else:
        raise HTTPException(status_code=400, detail="Loại dữ liệu không hợp lệ")


# ========== IMPORT EXCEL (YÊU CẦU KHÓA AN TOÀN) ==========

@router.post("/data/import/{data_type}")
async def import_data(data_type: str, payload: dict, request: Request):
    """Import danh mục từ Excel (yêu cầu khóa an toàn)"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    
    # Yêu cầu khóa an toàn
    safety_key = payload.get("safety_key", "")
    verify_safety_key(safety_key)
    
    wb = load_workbook_from_base64(payload)
    ws = wb.active
    conn = get_db()
    c = conn.cursor()
    
    try:
        if data_type == "products":
            return _import_products(c, ws)
        elif data_type == "customers":
            return _import_customers(c, ws)
        elif data_type == "suppliers":
            return _import_suppliers(c, ws)
        else:
            raise HTTPException(status_code=400, detail="Loại dữ liệu không hợp lệ")
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


def _import_products(c, ws):
    from app_api.products import ProductCreate
    created = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    # Build lookup maps for danh mục, đơn vị tính, kho
    c.execute("SELECT id, name FROM categories WHERE is_active=1")
    categories = {row['name'].strip().lower(): row['id'] for row in c.fetchall()}
    c.execute("SELECT id, name FROM units")
    units = {row['name'].strip().lower(): row['id'] for row in c.fetchall()}
    c.execute("SELECT id, name, code FROM warehouses WHERE is_active=1")
    warehouses = {row['name'].strip().lower(): row['id'] for row in c.fetchall()}
    warehouses.update({row['code'].strip().lower(): row['id'] for row in c.fetchall()})
    
    for idx, row in enumerate(rows, start=2):
        if not row or not any(v not in (None, "") for v in row):
            continue
        try:
            # Cột: Mã SP, Barcode, Tên hàng hóa, Danh mục, ĐVT, Kho, Giá vốn, Giá bán, Tồn kho ban đầu, Tồn tối thiểu, % Chiết khấu
            vals = list(row) + [None]*11
            code, barcode, name, category_name, unit_name, warehouse_name = vals[:6]
            cost_price, selling_price = vals[6], vals[7]
            initial_stock, min_stock, discount_rate = vals[8], vals[9], vals[10]
            
            if not name:
                raise ValueError("Thiếu tên sản phẩm")
            
            if not code:
                year = datetime.now().strftime("%Y")
                c.execute(f"SELECT MAX(CAST(SUBSTR(code, 4) AS INTEGER)) FROM products WHERE code LIKE 'SP{year}%'")
                max_num = c.fetchone()[0]
                next_num = (max_num or 0) + 1
                code = f"SP{year}{next_num:04d}"
            
            c.execute("SELECT id FROM products WHERE code=?", (code,))
            if c.fetchone():
                raise ValueError(f"Mã {code} đã tồn tại")
            if barcode:
                c.execute("SELECT id FROM products WHERE barcode=?", (barcode,))
                if c.fetchone():
                    raise ValueError(f"Barcode {barcode} đã tồn tại")
            
            # Lookup kho (bắt buộc)
            wh_key = str(warehouse_name or "").strip().lower()
            wh_id = warehouses.get(wh_key)
            if not wh_id:
                raise ValueError(f"Không tìm thấy kho '{warehouse_name}'. Vui lòng kiểm tra lại tên kho.")
            
            # Lookup danh mục (không bắt buộc)
            cat_id = None
            if category_name:
                cat_key = str(category_name).strip().lower()
                cat_id = categories.get(cat_key)
                if not cat_id:
                    errors.append({"row": idx, "error": f"Không tìm thấy danh mục '{category_name}', bỏ qua danh mục"})
            
            # Lookup đơn vị tính (không bắt buộc)
            unit_id = None
            if unit_name:
                unit_key = str(unit_name).strip().lower()
                unit_id = units.get(unit_key)
                if not unit_id:
                    errors.append({"row": idx, "error": f"Không tìm thấy ĐVT '{unit_name}', bỏ qua ĐVT"})
            
            c.execute("""
                INSERT INTO products (code, barcode, name, warehouse_id, category_id, unit_id, description, specifications, min_stock, max_stock, cost_price, selling_price, discount_rate, created_at, is_active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            """, (code, barcode, str(name).strip(), wh_id,
                  cat_id, unit_id,
                  None, None,
                  int(min_stock or 0), 0,
                  round(float(cost_price or 0), 2), round(float(selling_price or 0), 2),
                  round(float(discount_rate or 0), 2),
                  datetime.now().isoformat()))
            
            product_id = c.lastrowid
            init_stock = int(float(initial_stock or 0))
            c.execute("""
                INSERT OR IGNORE INTO inventory (product_id, warehouse_id, quantity_in_stock, updated_at)
                VALUES (?, ?, ?, ?)
            """, (product_id, wh_id, init_stock, datetime.now().isoformat()))
            created += 1
        except Exception as row_error:
            errors.append({"row": idx, "error": str(row_error)})
    
    c.connection.commit()
    return {"success": True, "created": created, "errors": errors, "message": f"Đã nhập {created} sản phẩm"}


def _import_customers(c, ws):
    created = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    for idx, row in enumerate(rows, start=2):
        if not row or not any(v not in (None, "") for v in row):
            continue
        try:
            name, contact_person, phone, email, address, city, tax_code = (list(row) + [None]*7)[:7]
            if not name:
                raise ValueError("Thiếu tên khách hàng")
            
            year = datetime.now().strftime("%Y")
            c.execute(f"SELECT MAX(CAST(SUBSTR(code, 7) AS INTEGER)) FROM customers WHERE code LIKE 'KH{year}%'")
            max_num = c.fetchone()[0]
            next_num = (max_num or 0) + 1
            code = f"KH{year}{next_num:04d}"
            
            c.execute("""
                INSERT INTO customers (code, name, contact_person, phone, email, address, city, tax_code, is_active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            """, (code, str(name).strip(), contact_person, phone, email, address, city, tax_code))
            created += 1
        except Exception as row_error:
            errors.append({"row": idx, "error": str(row_error)})
    
    c.connection.commit()
    return {"success": True, "created": created, "errors": errors, "message": f"Đã nhập {created} khách hàng"}


def _import_suppliers(c, ws):
    created = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    for idx, row in enumerate(rows, start=2):
        if not row or not any(v not in (None, "") for v in row):
            continue
        try:
            name, contact_person, phone, email, address, city, tax_code = (list(row) + [None]*7)[:7]
            if not name:
                raise ValueError("Thiếu tên nhà cung cấp")
            
            c.execute(f"SELECT MAX(CAST(SUBSTR(code, 4) AS INTEGER)) FROM suppliers WHERE code LIKE 'NCC%'")
            max_num = c.fetchone()[0]
            next_num = (max_num or 0) + 1
            code = f"NCC{next_num:03d}"
            
            c.execute("""
                INSERT INTO suppliers (code, name, contact_person, phone, email, address, city, tax_code, is_active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            """, (code, str(name).strip(), contact_person, phone, email, address, city, tax_code))
            created += 1
        except Exception as row_error:
            errors.append({"row": idx, "error": str(row_error)})
    
    c.connection.commit()
    return {"success": True, "created": created, "errors": errors, "message": f"Đã nhập {created} nhà cung cấp"}


# ========== XÓA DỮ LIỆU GIAO DỊCH (YÊU CẦU KHÓA AN TOÀN) ==========

class ClearTransactionsPayload(BaseModel):
    data_type: str  # 'import_orders', 'export_orders', 'all'
    confirm_text: str  # Must be "CONFIRM" to proceed
    safety_key: Optional[str] = None

@router.post("/data/clear-transactions")
async def clear_transactions(payload: ClearTransactionsPayload, request: Request):
    """Xóa dữ liệu giao dịch (nhập kho, xuất kho) - yêu cầu khóa an toàn"""
    _admin_only(request)
    
    # Yêu cầu khóa an toàn
    verify_safety_key(payload.safety_key)
    
    if payload.confirm_text != "CONFIRM":
        raise HTTPException(status_code=400, detail="Vui lòng nhập 'CONFIRM' để xác nhận xóa dữ liệu")
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        deleted_import = 0
        deleted_export = 0
        
        if payload.data_type in ("import_orders", "all"):
            # Delete import order items first
            c.execute("DELETE FROM import_order_items")
            deleted_import_items = c.rowcount
            # Delete import orders
            c.execute("DELETE FROM import_orders")
            deleted_import = c.rowcount
            # Delete related inventory transactions
            c.execute("DELETE FROM inventory_transactions WHERE transaction_type='import'")
        
        if payload.data_type in ("export_orders", "all"):
            # Delete export order items first
            c.execute("DELETE FROM export_order_items")
            deleted_export_items = c.rowcount
            # Delete export orders
            c.execute("DELETE FROM export_orders")
            deleted_export = c.rowcount
            # Delete related inventory transactions
            c.execute("DELETE FROM inventory_transactions WHERE transaction_type='export'")
        
        conn.commit()
        
        return {
            "success": True,
            "message": f"Đã xóa dữ liệu: {deleted_import} phiếu nhập, {deleted_export} phiếu xuất",
            "details": {
                "import_orders_deleted": deleted_import,
                "export_orders_deleted": deleted_export
            }
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi xóa dữ liệu: {str(e)}")
    finally:
        conn.close()


class ResetInventoryPayload(BaseModel):
    confirm_text: str  # Must be "CONFIRM" to proceed
    safety_key: Optional[str] = None

@router.post("/data/reset-inventory")
async def reset_inventory(payload: ResetInventoryPayload, request: Request):
    """Reset tồn kho về 0 - yêu cầu khóa an toàn"""
    _admin_only(request)
    
    # Yêu cầu khóa an toàn
    verify_safety_key(payload.safety_key)
    
    if payload.confirm_text != "CONFIRM":
        raise HTTPException(status_code=400, detail="Vui lòng nhập 'CONFIRM' để xác nhận reset tồn kho")
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        # Reset inventory
        c.execute("UPDATE inventory SET quantity_in_stock = 0, quantity_reserved = 0, quantity_available = 0, total_value = 0, updated_at = ?", 
                  (datetime.now().isoformat(),))
        updated_inventory = c.rowcount
        
        # Clear all inventory transactions
        c.execute("DELETE FROM inventory_transactions")
        deleted_transactions = c.rowcount
        
        conn.commit()
        
        return {
            "success": True,
            "message": f"Đã reset tồn kho: {updated_inventory} sản phẩm, xóa {deleted_transactions} giao dịch",
            "details": {
                "inventory_reset": updated_inventory,
                "transactions_deleted": deleted_transactions
            }
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi reset tồn kho: {str(e)}")
    finally:
        conn.close()


# ========== HELPERS ==========

def _format_size(size_bytes: int) -> str:
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"