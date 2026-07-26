"""
Shared Excel utility functions for WMS.
Consolidates duplicate helper functions from inventory.py, orders.py, products.py, partners.py, reports.py
"""
from datetime import datetime
from pathlib import Path
from urllib.parse import quote
import io
import base64
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell
from fastapi.responses import StreamingResponse
from database import get_db, get_company_profile

LOGO_PATH = Path(__file__).resolve().parents[1] / "static" / "logo.png"

MONEY_FORMAT = '#,##0'
MONEY_FORMAT_DEC = '#,##0.00'


def fmt_money(ws, row: int, col: int, value):
    """Write a monetary value with currency number format (no decimals)"""
    cell = ws.cell(row=row, column=col)
    cell.value = round(float(value or 0), 2)
    cell.number_format = MONEY_FORMAT
    return cell


def fmt_money_dec(ws, row: int, col: int, value):
    """Write a monetary value with currency number format (2 decimals)"""
    cell = ws.cell(row=row, column=col)
    cell.value = round(float(value or 0), 2)
    cell.number_format = MONEY_FORMAT_DEC
    return cell


def add_logo_and_brand(ws, title: str, subtitle: str = ""):
    """Add company logo and brand info to worksheet header"""
    profile = get_company_profile()
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.height = 54
            img.width = 54
            ws.add_image(img, "A2")
        except Exception:
            ws["A2"] = "LOGO"
            ws["A2"].font = Font(bold=True)
    else:
        ws["A2"] = "LOGO"
        ws["A2"].font = Font(bold=True)

    ws["C2"] = profile.get("short_name") or profile.get("company_name") or "An Tin Solution - WMS"
    line3 = subtitle or f"Ngay xuat: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["C3"] = f"{line3} | {profile.get('phone', '')}".strip(" |")
    ws["C2"].font = Font(bold=True, size=12)


def style_header(ws, row: int, last_col: int):
    """Style header row with fill and border"""
    fill = PatternFill("solid", fgColor="4472C4")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_body(ws, start_row: int, end_row: int, last_col: int):
    """Style body rows with border"""
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for row in range(start_row, end_row + 1):
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).border = border


def auto_fit_columns(ws, max_width: int = 38):
    """Auto-fit column widths based on content"""
    for column_cells in ws.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        first = next((cell for cell in column_cells if not isinstance(cell, MergedCell)), None)
        if not first:
            continue
        ws.column_dimensions[first.column_letter].width = min(max(len(v) for v in values) + 2, max_width)


def set_signatures(ws, row: int, delivery_label: str = "Người giao/nhận"):
    """Add signature lines at the bottom of worksheet"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="Người lập phiếu")
    ws.cell(row=row, column=4, value=delivery_label)
    ws.cell(row=row, column=7, value="Người duyệt")
    for col in (1, 4, 7):
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")


from app_api.date_utils import format_date_dd_mm_yyyy

def fmt_date(value):
    """Format date string to dd/mm/yyyy (UI/Excel)."""
    return format_date_dd_mm_yyyy(value)


def excel_response(wb, filename: str):
    """Convert workbook to StreamingResponse"""
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


def load_workbook_from_base64(payload: dict):
    """Load workbook from base64-encoded payload"""
    raw = payload.get("file_base64", "")
    if not raw:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Thiếu file_base64")
    if "," in raw:
        raw = raw.split(",", 1)[1]
    try:
        data = base64.b64decode(raw)
        return openpyxl.load_workbook(io.BytesIO(data))
    except Exception as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"File Excel không hợp lệ: {str(e)}")


def sample_wb(sheet_name: str, headers: list[str], example_row: list):
    """Create a sample workbook template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.append(example_row)
    fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    return wb


def set_report_brand(ws, title: str):
    """Add report brand header"""
    profile = get_company_profile()
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = profile.get("company_name") or ""
    ws["B2"] = profile.get("tax_code") or ""
    ws["D2"] = profile.get("phone") or ""
    ws["A3"] = profile.get("address") or ""
    ws["D3"] = profile.get("email") or ""


def style_header_row(ws, row_idx: int, start_col: int, end_col: int):
    """Style a header row (for reports)"""
    fill = PatternFill("solid", fgColor="E2E8F0")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill


def num(v):
    """Safe float conversion"""
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def build_order_detail_template(ws, order, items, profile, partner_info, is_import=True):
    """Build the detailed template for both import and export orders.
    is_import=True for import orders (phiếu nhập), False for export orders (phiếu xuất).
    Số dòng dữ liệu được tính động theo số lượng items.
    """
    partner_label = "Nhà cung cấp" if is_import else "Khách hàng"
    title = "PHIẾU NHẬP HÀNG" if is_import else "PHIẾU GIAO HÀNG"

    # Tính toán số dòng động
    num_items = len(items)
    header_row = 9
    start_row = 10
    end_data_row = start_row + num_items - 1  # Dòng cuối cùng của dữ liệu
    summary_notes_row = end_data_row + 1
    summary_row = summary_notes_row + 1
    footer_row = summary_row + 9
    signature_start_row = footer_row + 1
    total_rows = signature_start_row + 2  # 1 dòng nhãn + 1 dòng ghi chú + 1 dòng chữ ký

    ws.freeze_panes = "A10"
    ws.sheet_view.showGridLines = False

    # Header thông tin chung
    ws["A1"] = "NPP:"
    ws["C1"] = profile.get("company_name") or profile.get("short_name") or "An Tín Solution"
    ws["G1"] = "Số đơn hàng:"
    ws["I1"] = order.get("code", "")

    ws["A2"] = "Địa chỉ:"
    ws["C2"] = profile.get("address") or ""
    ws["G2"] = "Ngày đặt hàng:"
    ws["I2"] = fmt_date(order.get("order_date"))

    ws["A3"] = "Số điện thoại:"
    ws["C3"] = profile.get("phone") or ""
    if is_import:
        ws["G3"] = "Ngày nhập hàng:"
        ws["I3"] = fmt_date(order.get("received_date") or order.get("expected_date") or order.get("order_date"))
    else:
        ws["G3"] = "Ngày giao hàng:"
        ws["I3"] = fmt_date(order.get("expected_date") or order.get("shipped_date") or order.get("order_date"))

    ws["A4"] = "STK:"
    bank_text = " - ".join(filter(None, [profile.get("bank_account") or "", profile.get("bank_name") or ""]))
    ws["C4"] = bank_text
    ws["A5"] = title
    ws.merge_cells("A5:J5")
    ws["A5"].font = Font(size=16, bold=True)
    ws["A5"].alignment = Alignment(horizontal="center")

    # Thông tin đối tác
    ws["A6"] = f"Tên {partner_label.lower()}:"
    ws["C6"] = partner_info.get("name", "")
    ws["G6"] = "Địa chỉ:"
    if is_import:
        ws["I6"] = partner_info.get("address", "")
    else:
        ws["I6"] = order.get("shipping_address") or partner_info.get("address", "")

    ws["A7"] = "Điện thoại:"
    ws["C7"] = partner_info.get("phone", "")
    ws["G7"] = "Địa chỉ giao/nhận::"
    ws["I7"] = order.get("notes") or ""

    ws["A8"] = "NVBH:"
    ws["C8"] = order.get("created_by_name", "")
    ws["G8"] = "Điện thoại:"
    ws["I8"] = order.get("created_by_phone") or ""

    # Bảng dữ liệu hàng hóa A-J
    headers = [
        "STT", "Mã", "Tên hàng", "ĐVT", "SL",
        "Đơn giá", "Thành tiền", "% CK", "Số tiền giảm", "Thanh toán"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=header)
    style_header(ws, header_row, 10)

    # Data rows - động theo số lượng items
    subtotal = 0.0
    item_discount_total = 0.0
    for idx, item in enumerate(items):
        row = start_row + idx
        qty = num(item.get("quantity_ordered"))
        price = num(item.get("unit_price"))
        line_gross = qty * price
        line_discount = num(item.get("discount_rate"))
        discount_amount = line_gross * line_discount / 100.0
        line_net = num(item.get("total_price"))
        if not line_net:
            line_net = max(line_gross - discount_amount, 0)
        subtotal += line_net
        item_discount_total += discount_amount
        ws.cell(row=row, column=1, value=idx + 1)
        ws.cell(row=row, column=2, value=item.get("code"))
        ws.cell(row=row, column=3, value=item.get("product_name"))
        ws.cell(row=row, column=4, value=item.get("unit_name") or "")
        ws.cell(row=row, column=5, value=qty)
        fmt_money(ws, row, 6, price)
        fmt_money(ws, row, 7, line_gross)
        ws.cell(row=row, column=8, value=line_discount)
        fmt_money(ws, row, 9, discount_amount)
        fmt_money(ws, row, 10, line_net)

    style_body(ws, start_row, end_data_row, 10)

    # Summary area - vị trí động
    ws[f"A{summary_notes_row}"] = order.get("notes") or ""
    ws[f"F{summary_row}"] = "Giá trị Đơn hàng:"
    ws[f"J{summary_row}"] = f"=SUM(G{start_row}:G{end_data_row})"
    ws[f"J{summary_row}"].number_format = MONEY_FORMAT
    ws[f"F{summary_row + 1}"] = "Giảm giá theo mã hàng:"
    ws[f"J{summary_row + 1}"] = f"=SUM(I{start_row}:I{end_data_row})"
    ws[f"J{summary_row + 1}"].number_format = MONEY_FORMAT
    ws[f"F{summary_row + 2}"] = "Giảm giá trên đơn hàng:"
    fmt_money(ws, summary_row + 2, 10, order.get("discount_amount"))
    ws[f"F{summary_row + 3}"] = "Tiền phải trả:"
    ws[f"J{summary_row + 3}"] = f"=J{summary_row}-J{summary_row + 1}-J{summary_row + 2}"
    ws[f"J{summary_row + 3}"].number_format = MONEY_FORMAT
    ws[f"F{summary_row + 4}"] = "Đã thanh toán:"
    fmt_money(ws, summary_row + 4, 10, order.get("paid_amount"))
    ws[f"F{summary_row + 5}"] = "Công nợ còn lại:"
    ws[f"J{summary_row + 5}"] = f"=J{summary_row + 3}-J{summary_row + 4}"
    ws[f"J{summary_row + 5}"].number_format = MONEY_FORMAT
    ws[f"A{summary_notes_row}"].alignment = Alignment(wrap_text=True)
    for cell_ref in (f"F{summary_row}", f"F{summary_row + 1}", f"F{summary_row + 2}", f"F{summary_row + 3}", f"F{summary_row + 4}", f"F{summary_row + 5}"):
        ws[cell_ref].font = Font(bold=True)

    # Signature area - Khu vực chữ ký
    ws[f"A{footer_row}"] = "Quý khách vui lòng kiểm tra phiếu khi nhận hàng"
    if is_import:  # Phiếu nhập
        ws.merge_cells(start_row=signature_start_row, start_column=1, end_row=signature_start_row, end_column=2)
        ws.cell(row=signature_start_row, column=1, value="Nhà cung cấp")
        ws.merge_cells(start_row=signature_start_row, start_column=5, end_row=signature_start_row, end_column=6)
        ws.cell(row=signature_start_row, column=5, value="Người lập phiếu")
        ws.merge_cells(start_row=signature_start_row, start_column=9, end_row=signature_start_row, end_column=10)
        ws.cell(row=signature_start_row, column=9, value="Thủ kho")
    else:  # Phiếu xuất
        ws.merge_cells(start_row=signature_start_row, start_column=1, end_row=signature_start_row, end_column=2)
        ws.cell(row=signature_start_row, column=1, value="Khách hàng")
        ws.merge_cells(start_row=signature_start_row, start_column=5, end_row=signature_start_row, end_column=6)
        ws.cell(row=signature_start_row, column=5, value="NVGH")
        ws.merge_cells(start_row=signature_start_row, start_column=9, end_row=signature_start_row, end_column=10)
        ws.cell(row=signature_start_row, column=9, value="Thủ kho")

    for col in (1, 5, 9):
        cell = ws.cell(row=signature_start_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Row heights
    for row in range(1, total_rows + 1):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[footer_row].height = 22
    ws.row_dimensions[signature_start_row].height = 24
    ws.row_dimensions[signature_start_row + 1].height = 22
    ws.row_dimensions[signature_start_row + 2].height = 24

    # Column widths
    for col, width in {
        "A": 8, "B": 12, "C": 28, "D": 12, "E": 10,
        "F": 12, "G": 14, "H": 10, "I": 14, "J": 14
    }.items():
        ws.column_dimensions[col].width = width

    # Borders for data table only (header + data rows)
    style_header(ws, header_row, 10)
    style_body(ws, start_row, end_data_row, 10)

    # Number alignment
    for row in range(start_row, summary_row):
        for col in (5, 6, 7, 8, 9, 10):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")

    # Page setup
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5