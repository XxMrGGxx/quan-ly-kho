from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from datetime import datetime
from database import get_db
import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse
from app_api.debt import _calculate_supplier_debt as _calc_supplier_debt, _sync_supplier_debt as _sync_supplier_debt

router = APIRouter()


def _excel_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


def _load_workbook_from_base64(payload: dict):
    raw = payload.get("file_base64", "")
    if "," in raw:
        raw = raw.split(",", 1)[1]
    try:
        return openpyxl.load_workbook(io.BytesIO(base64.b64decode(raw)))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File Excel không hợp lệ: {str(e)}")


def _sample_wb(sheet_name: str, headers: list[str], example_row: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.append(example_row)
    fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    return wb

class PartnerCreate(BaseModel):
    name: str
    contact_person: str = None
    phone: str = None
    email: str = None
    address: str = None
    city: str = None
    tax_code: str = None

# --- KHÁCH HÀNG ---
@router.get("/customers")
async def list_customers(search: str = ""):
    conn = get_db()
    c = conn.cursor()
    query = "SELECT * FROM customers WHERE is_active=1"
    params = []
    if search and search.strip():
        query += " AND (name LIKE ? OR code LIKE ? OR phone LIKE ? OR email LIKE ?)"
        term = f"%{search}%"
        params.extend([term, term, term, term])
    query += " ORDER BY name LIMIT 200"
    c.execute(query, params)
    items = [dict(row) for row in c.fetchall()]
    conn.close()
    return {"items": items, "total": len(items)}

@router.post("/customers")
async def create_customer(data: PartnerCreate):
    conn = get_db()
    c = conn.cursor()
    
    # Auto Code
    year = datetime.now().strftime("%Y")
    # Use MAX() to get the highest existing code number, then increment
    # This handles gaps in sequence correctly
    # Code format: KH + YYYY + NNNN (e.g., KH20260001)
    c.execute(f"SELECT MAX(CAST(SUBSTR(code, 7) AS INTEGER)) FROM customers WHERE code LIKE 'KH{year}%'")
    max_num = c.fetchone()[0]
    next_num = (max_num or 0) + 1
    code = f"KH{year}{next_num:04d}"
    
    try:
        c.execute("""
            INSERT INTO customers (code, name, contact_person, phone, email, address, city, tax_code, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
        """, (code, data.name, data.contact_person, data.phone, data.email, data.address, data.city, data.tax_code))
        conn.commit()
        return {"id": c.lastrowid, "code": code}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


# --- NHÀ CUNG CẤP ---
@router.get("/suppliers")
async def list_suppliers(search: str = ""):
    conn = get_db()
    c = conn.cursor()
    query = "SELECT * FROM suppliers WHERE is_active=1"
    params = []
    if search and search.strip():
        query += " AND (name LIKE ? OR code LIKE ? OR phone LIKE ? OR email LIKE ?)"
        term = f"%{search}%"
        params.extend([term, term, term, term])
    query += " ORDER BY name LIMIT 200"
    c.execute(query, params)
    items = [dict(row) for row in c.fetchall()]
    
    # Tính lại công nợ thực tế từ import_orders và đồng bộ
    for item in items:
        supplier_id = item['id']
        # Tính công nợ thực tế từ phiếu nhập
        actual_debt = _calc_supplier_debt(conn, supplier_id)
        # Cập nhật current_debt trong DB
        c.execute("UPDATE suppliers SET current_debt = ? WHERE id = ?", (actual_debt, supplier_id))
        item['current_debt'] = actual_debt
    
    conn.commit()
    conn.close()
    return {"items": items, "total": len(items)}

@router.post("/suppliers")
async def create_supplier(data: PartnerCreate):
    conn = get_db()
    c = conn.cursor()
    
    # Use MAX() to get the highest existing code number, then increment
    # This handles gaps in sequence correctly
    # Code format: NCC + NNN (e.g., NCC001)
    c.execute(f"SELECT MAX(CAST(SUBSTR(code, 4) AS INTEGER)) FROM suppliers WHERE code LIKE 'NCC%'")
    max_num = c.fetchone()[0]
    next_num = (max_num or 0) + 1
    code = f"NCC{next_num:03d}"
    
    try:
        c.execute("""
            INSERT INTO suppliers (code, name, contact_person, phone, email, address, city, tax_code, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
        """, (code, data.name, data.contact_person, data.phone, data.email, data.address, data.city, data.tax_code))
        conn.commit()
        return {"id": c.lastrowid, "code": code}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@router.put("/customers/{customer_id}")
async def update_customer(customer_id: int, data: PartnerCreate):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            UPDATE customers 
            SET name=?, contact_person=?, phone=?, email=?, address=?, city=?, tax_code=?
            WHERE id=?
        """, (data.name, data.contact_person, data.phone, data.email, 
              data.address, data.city, data.tax_code, customer_id))
        conn.commit()
        return {"message": "Cập nhật thành công"}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE customers SET is_active=0 WHERE id=?", (customer_id,))
    conn.commit()
    conn.close()
    return {"message": "Đã xóa khách hàng"}

@router.put("/suppliers/{supplier_id}")
async def update_supplier(supplier_id: int, data: PartnerCreate):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            UPDATE suppliers 
            SET name=?, contact_person=?, phone=?, email=?, address=?, city=?, tax_code=?
            WHERE id=?
        """, (data.name, data.contact_person, data.phone, data.email,
              data.address, data.city, data.tax_code, supplier_id))
        conn.commit()
        return {"message": "Cập nhật thành công"}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE suppliers SET is_active=0 WHERE id=?", (supplier_id,))
    conn.commit()
    conn.close()
    return {"message": "Đã xóa nhà cung cấp"}
